import json
import os
import io
import boto3
import psycopg2
import openpyxl
import urllib.request
import urllib.parse
from datetime import datetime, timedelta
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SCHEMA = 't_p31606708_tech_buying_service'
HEADERS = {
    'Access-Control-Allow-Origin': '*',
    'Access-Control-Allow-Methods': 'GET, POST, OPTIONS',
    'Access-Control-Allow-Headers': 'Content-Type, X-Admin-Token, X-Employee-Token',
}

YM_HEADERS = ['Название', 'Цена', 'Описание', 'Категория', 'URL', 'Фото']

S3_KEY_CATALOG     = 'exports/catalog_export.xlsx'
S3_KEY_GOODS       = 'exports/goods_export.xlsx'
S3_KEY_TOOLS_FINAL = 'exports/tools_export.xlsx'

COL_WIDTHS = [50, 14, 80, 30, 60, 60]


def get_conn():
    return psycopg2.connect(os.environ['DATABASE_URL'])


def get_s3():
    return boto3.client(
        's3',
        endpoint_url='https://bucket.poehali.dev',
        aws_access_key_id=os.environ['AWS_ACCESS_KEY_ID'],
        aws_secret_access_key=os.environ['AWS_SECRET_ACCESS_KEY'],
    )


def check_auth(event):
    hdrs = event.get('headers', {})
    token = hdrs.get('X-Admin-Token') or hdrs.get('X-Employee-Token')
    if not token:
        return False
    # Фиксированный токен владельца
    if token == 'Mark2015N':
        return True
    # Проверка по токену сотрудника из БД
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        f"SELECT id FROM {SCHEMA}.employees WHERE auth_token='{token}' AND token_expires_at>NOW() AND is_active=true"
    )
    row = cur.fetchone()
    cur.close(); conn.close()
    return row is not None


def ok(data):
    return {'statusCode': 200, 'headers': HEADERS, 'body': json.dumps(data, ensure_ascii=False)}


def err(code, msg):
    return {'statusCode': code, 'headers': HEADERS, 'body': json.dumps({'error': msg}, ensure_ascii=False)}


def style_header(ws, fill_color):
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    font = Font(bold=True, color='FFFFFF', size=10)
    for col in range(1, len(YM_HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = COL_WIDTHS[col - 1]
    ws.row_dimensions[1].height = 30


def make_wb(sheet_name, fill_color, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(YM_HEADERS)
    style_header(ws, fill_color)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = 'A2'
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def clean_id(val):
    return str(val).strip().lstrip('\ufeff').lstrip('\u200b').strip()


def build_catalog(conn, model_photos=None, category_photos=None):
    model_photos = model_photos or {}
    category_photos = category_photos or {}
    cur = conn.cursor()
    cur.execute(f"""
        SELECT id, category, brand, model, color, storage,
               availability, price, description, photo_url
        FROM {SCHEMA}.catalog WHERE is_active = true
        ORDER BY category, brand, model
    """)
    rows = []
    for r in cur.fetchall():
        item_id, category, brand, model, color, storage, availability, price, description, photo_url = r
        name = ' '.join(x for x in [brand, model, storage, color] if x)
        short = ' '.join(x for x in [brand, model, storage] if x)
        desc = description or short
        if not price:
            continue
        photo = photo_url or model_photos.get(model) or category_photos.get(category) or ''
        rows.append([name, price, desc, category, photo, photo])
    cur.close()
    return make_wb('Каталог электроники', '1565C8', rows)


def build_goods(conn, model_photos=None, category_photos=None):
    model_photos = model_photos or {}
    category_photos = category_photos or {}
    cur = conn.cursor()
    cur.execute(f"""
        SELECT id, title, category, brand, model, condition,
               color, storage, sell_price, description, photo_url
        FROM {SCHEMA}.goods WHERE status = 'available'
        ORDER BY added_at DESC
    """)
    rows = []
    for r in cur.fetchall():
        item_id, title, category, brand, model, condition, color, storage, sell_price, description, photo_url = r
        name = title or ' '.join(x for x in [brand, model, storage, color] if x)
        short = ' '.join(x for x in [brand, model, storage, condition] if x)
        desc = description or short
        if not sell_price:
            continue
        photo = photo_url or model_photos.get(model) or category_photos.get(category) or ''
        rows.append([name, sell_price, desc, category, photo, photo])
    cur.close()
    return make_wb('Товары на складе', '27AE60', rows)


def build_tools(conn):
    cur = conn.cursor()
    # Только товары "В наличии" — лимит Яндекс Бизнес 10 000
    cur.execute(f"""
        SELECT article, name, brand, category,
               my_price, base_price, amount, image_url
        FROM {SCHEMA}.tools_products
        WHERE amount ILIKE '%наличи%'
        ORDER BY category, name
        LIMIT 9900
    """)
    rows = []
    for r in cur.fetchall():
        article, name, brand, category, my_price, base_price, amount, image_url = r
        price = float(my_price) if my_price and float(my_price) > 0 else (float(base_price) if base_price else '')
        short = ' '.join(x for x in [brand, name] if x)
        if not price:
            continue
        rows.append([name, price, short, category or '', image_url or '', image_url or ''])
    cur.close()
    return make_wb('Инструменты', 'E67E22', rows)


def cdn(key):
    key_id = os.environ['AWS_ACCESS_KEY_ID']
    return f"https://cdn.poehali.dev/projects/{key_id}/bucket/{key}"


def s3_exists(s3, key):
    try:
        s3.head_object(Bucket='files', Key=key)
        return True
    except Exception:
        return False


def put_xlsx(s3, key, data, filename):
    s3.put_object(
        Bucket='files', Key=key, Body=data,
        ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        ContentDisposition=f'attachment; filename="{filename}"',
    )


def handler(event: dict, context) -> dict:
    """Экспорт XLSX Яндекс Бизнес: 3 отдельных файла — каталог, товары, инструменты"""
    if event.get('httpMethod') == 'OPTIONS':
        return {'statusCode': 200, 'headers': HEADERS, 'body': ''}

    if not check_auth(event):
        return err(401, 'Unauthorized')

    method = event.get('httpMethod', 'GET')
    body = {}
    if method == 'POST':
        try:
            body = json.loads(event.get('body') or '{}')
        except Exception:
            body = {}

    qp = event.get('queryStringParameters') or {}
    action = body.get('action') if method == 'POST' else qp.get('action', 'status')
    s3 = get_s3()

    # GET — статус всех трёх файлов
    if action == 'status':
        return ok({
            'catalog_url': cdn(S3_KEY_CATALOG) if s3_exists(s3, S3_KEY_CATALOG) else None,
            'goods_url':   cdn(S3_KEY_GOODS)   if s3_exists(s3, S3_KEY_GOODS)   else None,
            'tools_url':   cdn(S3_KEY_TOOLS_FINAL) if s3_exists(s3, S3_KEY_TOOLS_FINAL) else None,
        })

    # POST action=catalog → catalog_export.xlsx
    if action == 'catalog':
        model_photos = body.get('model_photos', {})
        category_photos = body.get('category_photos', {})
        conn = get_conn()
        data = build_catalog(conn, model_photos, category_photos)
        conn.close()
        put_xlsx(s3, S3_KEY_CATALOG, data, 'catalog_export.xlsx')
        return ok({'status': 'catalog_done', 'catalog_url': cdn(S3_KEY_CATALOG)})

    # POST action=goods → goods_export.xlsx
    if action == 'goods':
        model_photos = body.get('model_photos', {})
        category_photos = body.get('category_photos', {})
        conn = get_conn()
        data = build_goods(conn, model_photos, category_photos)
        conn.close()
        put_xlsx(s3, S3_KEY_GOODS, data, 'goods_export.xlsx')
        return ok({'status': 'goods_done', 'goods_url': cdn(S3_KEY_GOODS)})

    # POST action=tools → tools_export.xlsx
    if action == 'tools':
        conn = get_conn()
        data = build_tools(conn)
        conn.close()
        put_xlsx(s3, S3_KEY_TOOLS_FINAL, data, 'tools_export.xlsx')
        return ok({'status': 'tools_done', 'tools_url': cdn(S3_KEY_TOOLS_FINAL)})

    # GET ?action=ym&days=14 → статистика Яндекс Метрики
    if action == 'ym':
        ym_token = os.environ.get('YM_TOKEN', '')
        if not ym_token:
            return err(503, 'YM_TOKEN not configured')
        days = int(qp.get('days', 14))
        date_to = datetime.today().strftime('%Y-%m-%d')
        date_from = (datetime.today() - timedelta(days=days - 1)).strftime('%Y-%m-%d')
        YM_ID = 108421419
        YM_API = 'https://api-metrika.yandex.net/stat/v1/data'

        def ym_get(params):
            qs = urllib.parse.urlencode(params)
            req = urllib.request.Request(f"{YM_API}?{qs}", headers={'Authorization': f'OAuth {ym_token}'})
            with urllib.request.urlopen(req, timeout=15) as r:
                return json.loads(r.read())

        daily = ym_get({'ids': YM_ID, 'metrics': 'ym:s:visits,ym:s:pageviews,ym:s:users,ym:s:bounceRate,ym:s:avgVisitDurationSeconds',
                        'dimensions': 'ym:s:date', 'date1': date_from, 'date2': date_to, 'sort': 'ym:s:date', 'limit': 100})
        stats = [{'date': r['dimensions'][0]['name'], 'visits': int(r['metrics'][0]), 'pageviews': int(r['metrics'][1]),
                  'users': int(r['metrics'][2]), 'bounceRate': round(r['metrics'][3], 1), 'avgDuration': int(r['metrics'][4])}
                 for r in daily.get('data', [])]

        src = ym_get({'ids': YM_ID, 'metrics': 'ym:s:visits', 'dimensions': 'ym:s:trafficSourceName',
                      'date1': date_from, 'date2': date_to, 'sort': '-ym:s:visits', 'limit': 10})
        tv = sum(r['metrics'][0] for r in src.get('data', [])) or 1
        sources = [{'source': r['dimensions'][0]['name'], 'visits': int(r['metrics'][0]), 'pct': round(int(r['metrics'][0]) / tv * 100)} for r in src.get('data', [])]

        dev = ym_get({'ids': YM_ID, 'metrics': 'ym:s:visits', 'dimensions': 'ym:s:deviceCategory',
                      'date1': date_from, 'date2': date_to, 'sort': '-ym:s:visits', 'limit': 5})
        td = sum(r['metrics'][0] for r in dev.get('data', [])) or 1
        DN = {'desktop': 'Компьютер', 'mobile': 'Смартфон', 'tablet': 'Планшет'}
        devices = [{'device': DN.get(r['dimensions'][0]['name'], r['dimensions'][0]['name']), 'visits': int(r['metrics'][0]),
                    'pct': round(int(r['metrics'][0]) / td * 100)} for r in dev.get('data', [])]

        geo_r = ym_get({'ids': YM_ID, 'metrics': 'ym:s:visits', 'dimensions': 'ym:s:regionCity',
                        'date1': date_from, 'date2': date_to, 'sort': '-ym:s:visits', 'limit': 10})
        tg2 = sum(r['metrics'][0] for r in geo_r.get('data', [])) or 1
        geo = [{'city': r['dimensions'][0].get('name') or 'Неизвестно', 'visits': int(r['metrics'][0]),
                'pct': round(int(r['metrics'][0]) / tg2 * 100)} for r in geo_r.get('data', [])]

        pg = ym_get({'ids': YM_ID, 'metrics': 'ym:pv:pageviews,ym:pv:avgTimeOnPage', 'dimensions': 'ym:pv:URLPathFull',
                     'date1': date_from, 'date2': date_to, 'sort': '-ym:pv:pageviews', 'limit': 10})
        pages = [{'url': r['dimensions'][0]['name'], 'views': int(r['metrics'][0]), 'avgTime': int(r['metrics'][1])} for r in pg.get('data', [])]

        return ok({'stats': stats, 'sources': sources, 'devices': devices, 'geo': geo, 'pages': pages})

    return err(400, 'Unknown action')