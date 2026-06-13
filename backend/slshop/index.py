import json
import os
import csv
import io
import re
import base64
import uuid
import urllib.request
import urllib.error
from datetime import datetime, timedelta
from decimal import Decimal

import psycopg2
import psycopg2.extras

HEADERS = {
    'Access-Control-Allow-Origin': '*',
    'Access-Control-Allow-Methods': 'GET, POST, PUT, DELETE, OPTIONS',
    'Access-Control-Allow-Headers': 'Content-Type, X-Employee-Token, Authorization',
    'Content-Type': 'application/json',
}
SCHEMA = 't_p31606708_tech_buying_service'

# URL пуш-хаба (см. backend/notify-push)
PUSH_URL = "https://functions.poehali.dev/0a041e7f-92ab-4dbf-86f0-cd09e3eabfbd"
MAX_BOT_URL = "https://functions.poehali.dev/4618b13e-cd61-4167-b943-0f3d439d0c8c"


def _notify_sale(title: str, item_name: str, amount: float, employee_name: str, qty: int = 1) -> None:
    """Уведомляет сотрудников в MAX и Telegram о продаже со смартломбарда."""
    try:
        qty_text = f' ({qty} шт)' if qty > 1 else ''
        text = (
            f"💰 *Продажа со склада*\n\n"
            f"📱 {item_name}{qty_text}\n"
            f"💵 *{int(amount):,} ₽*\n".replace(',', '\u00a0') +
            f"👨‍💼 {employee_name}"
        )
        data = json.dumps({'text': text}).encode('utf-8')
        req = urllib.request.Request(
            f'{MAX_BOT_URL}?action=staff_send', data=data, method='POST',
            headers={'Content-Type': 'application/json'},
        )
        urllib.request.urlopen(req, timeout=5).read()
    except Exception as e:
        print(f'[slshop][notify_sale MAX] {e}')
    try:
        tg_token = os.environ.get('TELEGRAM_BOT_TOKEN', '')
        tg_chat = os.environ.get('TELEGRAM_CHAT_ID', '')
        if tg_token and tg_chat:
            qty_text = f' ({qty} шт)' if qty > 1 else ''
            text = (
                f"💰 *Продажа со склада*\n\n"
                f"📱 {item_name}{qty_text}\n"
                f"💵 *{int(amount):,} ₽*\n".replace(',', '\u00a0') +
                f"👨‍💼 {employee_name}"
            )
            data = json.dumps({'chat_id': tg_chat, 'text': text, 'parse_mode': 'Markdown'}).encode('utf-8')
            req = urllib.request.Request(
                f'https://api.telegram.org/bot{tg_token}/sendMessage', data=data, method='POST',
                headers={'Content-Type': 'application/json'},
            )
            urllib.request.urlopen(req, timeout=5).read()
    except Exception as e:
        print(f'[slshop][notify_sale TG] {e}')


def _send_push_event(title: str, body: str, url: str = "/staff", tag: str = "slshop") -> None:
    """Шлёт push сотрудникам через notify-push. Тихо игнорирует ошибки."""
    try:
        admin = os.environ.get('ADMIN_TOKEN', '')
        if not admin:
            return
        data = json.dumps({'title': title, 'body': body, 'url': url, 'tag': tag}).encode('utf-8')
        req = urllib.request.Request(
            PUSH_URL, data=data, method='POST',
            headers={'X-Service-Token': admin, 'Content-Type': 'application/json'},
        )
        urllib.request.urlopen(req, timeout=3).read()
    except Exception:
        pass


def get_conn():
    return psycopg2.connect(os.environ['DATABASE_URL'])


def _ok(data, status=200):
    return {'statusCode': status, 'headers': HEADERS, 'body': json.dumps(data, ensure_ascii=False, default=_json_default)}


def _err(status, msg):
    return {'statusCode': status, 'headers': HEADERS, 'body': json.dumps({'error': msg}, ensure_ascii=False)}


def _json_default(o):
    if isinstance(o, (datetime,)):
        return o.isoformat()
    if isinstance(o, Decimal):
        return float(o)
    if hasattr(o, 'isoformat'):
        return o.isoformat()
    return str(o)


def get_employee_by_token(token: str):
    if not token:
        return None
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT id, full_name, login, role FROM {SCHEMA}.employees WHERE auth_token=%s AND token_expires_at>NOW() AND is_active=true", (token,))
    row = cur.fetchone(); cur.close(); conn.close()
    return {'id': row[0], 'full_name': row[1], 'login': row[2], 'role': row[3]} if row else None


def _esc(v):
    if v is None:
        return 'NULL'
    if isinstance(v, bool):
        return 'TRUE' if v else 'FALSE'
    if isinstance(v, (int, float, Decimal)):
        return str(v)
    s = str(v).replace("'", "''")
    return f"'{s}'"


def _norm_key(s: str) -> str:
    if not s:
        return ''
    s = s.lower().strip()
    s = re.sub(r'[^\w\s]+', ' ', s, flags=re.UNICODE)
    s = re.sub(r'\s+', ' ', s)
    return s.strip()


# ============ Categories ============
def list_categories():
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        f"SELECT id, name, slug, icon, color, sort_order, is_active, parent_id, depth, path "
        f"FROM {SCHEMA}.slshop_categories WHERE is_active=true "
        f"ORDER BY COALESCE(parent_id, id), depth, sort_order, name"
    )
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [dict(r) for r in rows]


def create_category(body):
    name = (body.get('name') or '').strip()
    slug = (body.get('slug') or '').strip() or _norm_key(name).replace(' ', '_')
    icon = body.get('icon') or 'Package'
    sort_order = int(body.get('sort_order') or 100)
    parent_id = body.get('parent_id')
    if not name:
        return _err(400, 'Имя обязательно')
    depth = 0
    path = name
    if parent_id:
        conn = get_conn(); cur = conn.cursor()
        cur.execute(f"SELECT name, depth FROM {SCHEMA}.slshop_categories WHERE id={int(parent_id)}")
        row = cur.fetchone()
        if row:
            depth = (row[1] or 0) + 1
            path = (row[0] or '') + ' / ' + name
        cur.close(); conn.close()
    conn = get_conn(); cur = conn.cursor()
    sql = (
        f"INSERT INTO {SCHEMA}.slshop_categories (name, slug, icon, sort_order, parent_id, depth, path) "
        f"VALUES ({_esc(name)}, {_esc(slug)}, {_esc(icon)}, {sort_order}, {_esc(parent_id)}, {depth}, {_esc(path)}) RETURNING id"
    )
    cur.execute(sql)
    nid = cur.fetchone()[0]
    conn.commit(); cur.close(); conn.close()
    return _ok({'id': nid})


# ============ Discount rules ============
def list_discount_rules():
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        f"SELECT r.*, c.name AS category_name FROM {SCHEMA}.slshop_discount_rules r "
        f"LEFT JOIN {SCHEMA}.slshop_categories c ON c.id=r.category_id ORDER BY r.id DESC"
    )
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [dict(r) for r in rows]


def save_discount_rule(body):
    rid = body.get('id')
    fields = {
        'name': (body.get('name') or 'Уценка').strip(),
        'category_id': body.get('category_id') or None,
        'apply_to_all': bool(body.get('apply_to_all', False)),
        'period_days': int(body.get('period_days') or 30),
        'percent': float(body.get('percent') or 5),
        'use_market_price': bool(body.get('use_market_price', False)),
        'use_duplicates_dependency': bool(body.get('use_duplicates_dependency', False)),
        'rounding': body.get('rounding') or 'one_decimal',
        'is_active': bool(body.get('is_active', True)),
        'max_discount_percent': body.get('max_discount_percent'),
        'min_price': body.get('min_price'),
    }
    conn = get_conn(); cur = conn.cursor()
    if rid:
        sets = ', '.join([f"{k}={_esc(v)}" for k, v in fields.items()])
        cur.execute(f"UPDATE {SCHEMA}.slshop_discount_rules SET {sets}, updated_at=NOW() WHERE id={int(rid)} RETURNING id")
        row = cur.fetchone()
        nid = row[0] if row else rid
    else:
        cols = ', '.join(fields.keys())
        vals = ', '.join([_esc(v) for v in fields.values()])
        cur.execute(f"INSERT INTO {SCHEMA}.slshop_discount_rules ({cols}) VALUES ({vals}) RETURNING id")
        nid = cur.fetchone()[0]
    conn.commit(); cur.close(); conn.close()
    return _ok({'id': nid})


def discount_rule_toggle(body):
    rid = body.get('id')
    if not rid:
        return _err(400, 'id обязателен')
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"UPDATE {SCHEMA}.slshop_discount_rules SET is_active = NOT is_active WHERE id={int(rid)} RETURNING is_active")
    row = cur.fetchone()
    conn.commit(); cur.close(); conn.close()
    return _ok({'is_active': row[0] if row else None})


def _round_price(p, mode):
    if p is None:
        return p
    p = float(p)
    if mode == 'integer':
        return round(p)
    if mode == 'tens':
        return round(p / 10) * 10
    if mode == 'fifty':
        return round(p / 50) * 50
    if mode == 'hundred':
        return round(p / 100) * 100
    return round(p, 1)


def discount_rule_apply(body):
    """Применяет правило уценки ко всем подходящим товарам один раз. Возвращает список изменений."""
    rid = body.get('id')
    if not rid:
        return _err(400, 'id обязателен')
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(f"SELECT * FROM {SCHEMA}.slshop_discount_rules WHERE id={int(rid)}")
    rule = cur.fetchone()
    if not rule:
        cur.close(); conn.close(); return _err(404, 'Правило не найдено')
    rule = dict(rule)
    period = int(rule['period_days'])
    pct = float(rule['percent'])
    where = ["status IN ('stock','showcase','consignment')", "sell_price > 0"]
    if not rule['apply_to_all'] and rule['category_id']:
        where.append(f"category_id={int(rule['category_id'])}")
    where.append(
        f"(last_discount_at IS NULL AND buy_at IS NOT NULL AND buy_at < NOW() - INTERVAL '{period} days' "
        f"OR last_discount_at IS NOT NULL AND last_discount_at < NOW() - INTERVAL '{period} days')"
    )
    if rule.get('min_price'):
        where.append(f"sell_price > {float(rule['min_price'])}")
    sql = f"SELECT id, sell_price, original_sell_price, discount_count FROM {SCHEMA}.slshop_items WHERE " + ' AND '.join(where) + ' LIMIT 500'
    cur.execute(sql)
    rows = cur.fetchall()
    applied = 0
    log = []
    for r in rows:
        old = float(r['sell_price'])
        new_price = old * (1 - pct / 100.0)
        new_price = _round_price(new_price, rule.get('rounding'))
        if rule.get('min_price') and new_price < float(rule['min_price']):
            new_price = float(rule['min_price'])
        # max_discount_percent — общая просадка от исходной цены
        orig = r['original_sell_price'] or old
        if rule.get('max_discount_percent'):
            min_allowed = float(orig) * (1 - float(rule['max_discount_percent']) / 100.0)
            if new_price < min_allowed:
                new_price = _round_price(min_allowed, rule.get('rounding'))
        if abs(new_price - old) < 0.01:
            continue
        cur.execute(
            f"UPDATE {SCHEMA}.slshop_items SET sell_price={new_price}, "
            f"original_sell_price=COALESCE(original_sell_price, {old}), "
            f"last_discount_at=NOW(), discount_count=COALESCE(discount_count,0)+1 WHERE id={int(r['id'])}"
        )
        cur.execute(
            f"INSERT INTO {SCHEMA}.slshop_discount_log (item_id, rule_id, price_before, price_after) "
            f"VALUES ({int(r['id'])}, {int(rid)}, {old}, {new_price})"
        )
        log.append({'item_id': r['id'], 'old': old, 'new': new_price})
        applied += 1
    conn.commit(); cur.close(); conn.close()
    return _ok({'applied': applied, 'log': log})


# ============ Revisions ============
def list_revisions():
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(f"SELECT * FROM {SCHEMA}.slshop_revisions ORDER BY id DESC LIMIT 50")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [dict(r) for r in rows]


def revision_create(body, employee):
    name = (body.get('name') or f'Ревизия {datetime.now().strftime("%d.%m.%Y %H:%M")}').strip()
    category_id = body.get('category_id')
    scope_status = body.get('scope_status') or 'stock'
    conn = get_conn(); cur = conn.cursor()
    # Валидация category_id
    ok_cat, cat_norm, _e = _validate_fk(cur, 'slshop_categories', category_id, 'category_id')
    category_id = cat_norm if ok_cat else None
    where = [f"status={_esc(scope_status)}"]
    if category_id:
        where.append(f"category_id={int(category_id)}")
    wsql = ' AND '.join(where)
    cur.execute(
        f"INSERT INTO {SCHEMA}.slshop_revisions (name, category_id, scope_status, started_by) "
        f"VALUES ({_esc(name)}, {_esc(category_id)}, {_esc(scope_status)}, {_esc(employee.get('full_name'))}) RETURNING id"
    )
    rev_id = cur.fetchone()[0]
    cur.execute(f"SELECT id FROM {SCHEMA}.slshop_items WHERE {wsql}")
    item_ids = [r[0] for r in cur.fetchall()]
    for iid in item_ids:
        cur.execute(
            f"INSERT INTO {SCHEMA}.slshop_revision_items (revision_id, item_id, state) "
            f"VALUES ({rev_id}, {iid}, 'pending')"
        )
    cur.execute(f"UPDATE {SCHEMA}.slshop_revisions SET total_expected={len(item_ids)} WHERE id={rev_id}")
    conn.commit(); cur.close(); conn.close()
    return _ok({'id': rev_id, 'total': len(item_ids)})


def revision_get(params):
    rid = params.get('id')
    if not rid:
        return _err(400, 'id обязателен')
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(f"SELECT * FROM {SCHEMA}.slshop_revisions WHERE id={int(rid)}")
    rev = cur.fetchone()
    if not rev:
        cur.close(); conn.close(); return _err(404, 'Ревизия не найдена')
    cur.execute(
        f"SELECT ri.*, i.title, i.imei, i.sku, i.sell_price FROM {SCHEMA}.slshop_revision_items ri "
        f"LEFT JOIN {SCHEMA}.slshop_items i ON i.id=ri.item_id "
        f"WHERE ri.revision_id={int(rid)} ORDER BY ri.state, i.title"
    )
    items = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return _ok({'revision': dict(rev), 'items': items})


def revision_scan(body, employee):
    rev_id = body.get('revision_id')
    code = (body.get('code') or '').strip()
    if not rev_id or not code:
        return _err(400, 'revision_id и code обязательны')
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    # ищем товар по sku или imei
    cur.execute(
        f"SELECT id, title FROM {SCHEMA}.slshop_items WHERE sku={_esc(code)} OR imei={_esc(code)} OR external_id={_esc(code)} LIMIT 1"
    )
    item = cur.fetchone()
    if not item:
        # помечаем как extra
        cur.execute(
            f"INSERT INTO {SCHEMA}.slshop_revision_items (revision_id, scanned_code, state, scanned_at, scanned_by) "
            f"VALUES ({int(rev_id)}, {_esc(code)}, 'extra', NOW(), {_esc(employee.get('full_name'))})"
        )
        conn.commit(); cur.close(); conn.close()
        return _ok({'state': 'extra', 'message': 'Товар не найден в базе — отмечен как лишний'})
    iid = item['id']
    cur.execute(
        f"SELECT id, state FROM {SCHEMA}.slshop_revision_items WHERE revision_id={int(rev_id)} AND item_id={iid} LIMIT 1"
    )
    ri = cur.fetchone()
    if ri:
        cur.execute(
            f"UPDATE {SCHEMA}.slshop_revision_items SET state='found', scanned_at=NOW(), scanned_by={_esc(employee.get('full_name'))}, scanned_code={_esc(code)} WHERE id={ri['id']}"
        )
        state = 'found'
    else:
        cur.execute(
            f"INSERT INTO {SCHEMA}.slshop_revision_items (revision_id, item_id, scanned_code, state, scanned_at, scanned_by) "
            f"VALUES ({int(rev_id)}, {iid}, {_esc(code)}, 'extra', NOW(), {_esc(employee.get('full_name'))})"
        )
        state = 'extra'
    conn.commit(); cur.close(); conn.close()
    return _ok({'state': state, 'item_id': iid, 'title': item['title']})


def list_branches():
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(f"SELECT * FROM {SCHEMA}.slshop_branches WHERE is_active=true ORDER BY sort_order, name")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [dict(r) for r in rows]


def items_move_branch(body, employee):
    """Массовый/одиночный перенос товаров на другой филиал."""
    if not employee:
        return _err(401, 'Не авторизован')
    item_ids = body.get('item_ids') or []
    move_all = bool(body.get('move_all', False))
    branch_id = body.get('branch_id')
    if not branch_id:
        return _err(400, 'branch_id обязателен')
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT id, name FROM {SCHEMA}.slshop_branches WHERE id={int(branch_id)}")
    br = cur.fetchone()
    if not br:
        cur.close(); conn.close(); return _err(404, 'Филиал не найден')
    branch_name = br[1]
    if move_all:
        cur.execute(
            f"UPDATE {SCHEMA}.slshop_items SET branch_id={int(branch_id)}, updated_at=NOW() "
            f"WHERE status IN ('stock','showcase','consignment')"
        )
        moved = cur.rowcount
    else:
        if not item_ids or not isinstance(item_ids, list):
            cur.close(); conn.close(); return _err(400, 'item_ids должен быть непустым массивом')
        ids_int = [int(x) for x in item_ids if x]
        if not ids_int:
            cur.close(); conn.close(); return _err(400, 'нет валидных id')
        ids_csv = ','.join(str(x) for x in ids_int)
        cur.execute(
            f"UPDATE {SCHEMA}.slshop_items SET branch_id={int(branch_id)}, updated_at=NOW() "
            f"WHERE id IN ({ids_csv})"
        )
        moved = cur.rowcount
    _log_event(cur, 'move', 'items', None, f'Перенос {moved} товаров → {branch_name}',
               'Массовый перенос' if move_all else f'Выбрано {len(item_ids)} шт',
               None, int(branch_id), employee)
    conn.commit(); cur.close(); conn.close()
    return _ok({'moved': moved, 'branch_id': int(branch_id), 'branch_name': branch_name})


def client_passport_upload(body, employee):
    """Загрузка фото паспорта клиента (base64 → S3)."""
    if not employee:
        return _err(401, 'Не авторизован')
    image_b64 = body.get('image_base64') or ''
    field = (body.get('field') or 'passport_photo_url').strip()
    if field not in ('passport_photo_url', 'passport_photo2_url', 'face_photo_url'):
        return _err(400, 'field должен быть passport_photo_url|passport_photo2_url|face_photo_url')
    client_id = body.get('client_id')
    if not image_b64:
        return _err(400, 'image_base64 обязателен')
    try:
        import boto3 as _boto3
        import base64 as _b64
        if ',' in image_b64:
            image_b64 = image_b64.split(',', 1)[1]
        raw = _b64.b64decode(image_b64)
    except Exception as e:
        return _err(400, f'Ошибка декодирования: {e}')
    if len(raw) > 10 * 1024 * 1024:
        return _err(400, 'Файл больше 10 МБ')
    s3 = _boto3.client(
        's3', endpoint_url='https://bucket.poehali.dev',
        aws_access_key_id=os.environ['AWS_ACCESS_KEY_ID'],
        aws_secret_access_key=os.environ['AWS_SECRET_ACCESS_KEY'],
    )
    import uuid as _uuid
    fname = f"slshop/passports/{_uuid.uuid4().hex}.jpg"
    s3.put_object(Bucket='files', Key=fname, Body=raw, ContentType='image/jpeg')
    cdn = f"https://cdn.poehali.dev/projects/{os.environ['AWS_ACCESS_KEY_ID']}/bucket/{fname}"
    if client_id:
        conn = get_conn(); cur = conn.cursor()
        cur.execute(f"UPDATE {SCHEMA}.slshop_clients SET {field}={_esc(cdn)}, updated_at=NOW() WHERE id={int(client_id)}")
        conn.commit(); cur.close(); conn.close()
    return _ok({'url': cdn, 'field': field})


def employee_set_role(body, employee):
    """Назначить/изменить роль сотрудника. Только владелец."""
    if not employee or employee.get('role') != 'owner':
        return _err(403, 'Назначать роли может только владелец')
    emp_id = body.get('employee_id')
    role = (body.get('role') or '').strip().lower()
    if not emp_id or not role:
        return _err(400, 'employee_id и role обязательны')
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT 1 FROM {SCHEMA}.slshop_roles WHERE code={_esc(role)} LIMIT 1")
    if not cur.fetchone():
        cur.close(); conn.close(); return _err(400, f'Роль "{role}" не найдена')
    cur.execute(f"UPDATE {SCHEMA}.employees SET role={_esc(role)} WHERE id={int(emp_id)} RETURNING id, full_name, role")
    row = cur.fetchone()
    if not row:
        conn.rollback(); cur.close(); conn.close()
        return _err(404, 'Сотрудник не найден')
    _log_event(cur, 'role_change', 'employee', int(emp_id), f'Смена роли: {row[1]} → {role}',
               None, None, None, employee)
    conn.commit(); cur.close(); conn.close()
    return _ok({'id': row[0], 'full_name': row[1], 'role': row[2]})


def employees_list_full(employee):
    """Список всех сотрудников с ролями."""
    if not employee or employee.get('role') not in ('owner', 'manager', 'admin'):
        return _err(403, 'Нет доступа')
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        f"SELECT e.id, e.full_name, e.login, e.role, e.is_active, e.created_at, e.last_seen_at, e.email, e.phone, "
        f"r.name AS role_name, r.description AS role_description "
        f"FROM {SCHEMA}.employees e "
        f"LEFT JOIN {SCHEMA}.slshop_roles r ON r.code=e.role "
        f"ORDER BY e.is_active DESC, e.id"
    )
    rows = cur.fetchall()
    cur.close(); conn.close()
    return _ok([dict(r) for r in rows])


def _log_event(cur, event_type, entity_type, entity_id, title, description=None, amount=None, branch_id=None, employee=None):
    """Лог в отдельной savepoint, чтобы при ошибке не ломать основную транзакцию."""
    try:
        emp_name = employee.get('full_name') if employee else None
        # Проверим что branch существует
        if branch_id:
            try:
                cur.execute(f"SELECT 1 FROM {SCHEMA}.slshop_branches WHERE id={int(branch_id)}")
                if not cur.fetchone():
                    branch_id = None
            except Exception:
                branch_id = None
        cur.execute(f"SAVEPOINT log_event")
        try:
            cur.execute(
                f"INSERT INTO {SCHEMA}.slshop_events (event_type, entity_type, entity_id, title, description, amount, branch_id, employee_name) "
                f"VALUES ({_esc(event_type)}, {_esc(entity_type)}, {_esc(entity_id)}, {_esc(title)}, {_esc(description)}, {_esc(amount)}, {_esc(branch_id)}, {_esc(emp_name)})"
            )
            cur.execute(f"RELEASE SAVEPOINT log_event")
        except Exception:
            try:
                cur.execute(f"ROLLBACK TO SAVEPOINT log_event")
            except Exception:
                pass
    except Exception:
        pass


def list_events(params):
    """Журнал всех событий с фильтрами"""
    event_type = (params.get('event_type') or '').strip()
    date_from = (params.get('date_from') or '').strip()
    date_to = (params.get('date_to') or '').strip()
    branch_id = params.get('branch_id')
    where = []
    if event_type:
        where.append(f"event_type={_esc(event_type)}")
    if date_from:
        where.append(f"created_at >= {_esc(date_from)}")
    if date_to:
        where.append(f"created_at < {_esc(date_to)}::date + INTERVAL '1 day'")
    if branch_id:
        where.append(f"branch_id={int(branch_id)}")
    wsql = (' WHERE ' + ' AND '.join(where)) if where else ''
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        f"SELECT e.*, b.name AS branch_name FROM {SCHEMA}.slshop_events e "
        f"LEFT JOIN {SCHEMA}.slshop_branches b ON b.id=e.branch_id "
        f"{wsql} ORDER BY e.created_at DESC LIMIT 500"
    )
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [dict(r) for r in rows]


def inventory_report(params):
    """Складской баланс и ДДС: остатки на начало/конец месяца, инвестиции vs вычеты,
    движение по дням и месяцам, прирост товарного запаса, ROI."""
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # ── 1. Итоговые цифры с начала бизнеса ──────────────────────────────────
    cur.execute(f"""
        SELECT
            MIN(created_at)                                                         AS started_at,
            COUNT(*)                                                                AS total_items,
            COUNT(*) FILTER (WHERE status IN ('stock','showcase','consignment'))    AS in_stock_count,
            COUNT(*) FILTER (WHERE status = 'sold')                                 AS sold_count,
            COUNT(*) FILTER (WHERE status = 'returned')                             AS returned_count,
            COALESCE(SUM(buy_price), 0)                                             AS total_invested,
            COALESCE(SUM(buy_price) FILTER (WHERE status IN ('stock','showcase','consignment')), 0) AS stock_value_buy,
            COALESCE(SUM(sell_price) FILTER (WHERE status IN ('stock','showcase','consignment')), 0) AS stock_value_sell,
            COALESCE(SUM(sell_price) FILTER (WHERE status = 'sold'), 0)             AS total_revenue,
            COALESCE(SUM(buy_price)  FILTER (WHERE status = 'sold'), 0)             AS total_cogs,
            COALESCE(SUM(sell_price - buy_price) FILTER (WHERE status = 'sold'), 0) AS total_profit
        FROM {SCHEMA}.slshop_items
    """)
    totals = dict(cur.fetchone())

    # ── 2. По месяцам: закупки и продажи ────────────────────────────────────
    cur.execute(f"""
        SELECT
            TO_CHAR(DATE_TRUNC('month', created_at), 'YYYY-MM') AS month,
            COUNT(*)                                             AS bought_count,
            COALESCE(SUM(buy_price), 0)                         AS bought_sum
        FROM {SCHEMA}.slshop_items
        WHERE created_at IS NOT NULL
        GROUP BY DATE_TRUNC('month', created_at)
        ORDER BY DATE_TRUNC('month', created_at)
    """)
    by_month_buy = [dict(r) for r in cur.fetchall()]

    cur.execute(f"""
        SELECT
            TO_CHAR(DATE_TRUNC('month', sell_at), 'YYYY-MM') AS month,
            COUNT(*)                                           AS sold_count,
            COALESCE(SUM(sell_price), 0)                      AS revenue,
            COALESCE(SUM(buy_price), 0)                       AS cogs,
            COALESCE(SUM(sell_price - buy_price), 0)          AS profit
        FROM {SCHEMA}.slshop_items
        WHERE sell_at IS NOT NULL AND status = 'sold'
        GROUP BY DATE_TRUNC('month', sell_at)
        ORDER BY DATE_TRUNC('month', sell_at)
    """)
    by_month_sell = [dict(r) for r in cur.fetchall()]

    # ── 3. По дням: закупки ─────────────────────────────────────────────────
    cur.execute(f"""
        SELECT
            DATE(buy_at)                    AS day,
            COUNT(*)                        AS bought_count,
            COALESCE(SUM(buy_price), 0)     AS bought_sum
        FROM {SCHEMA}.slshop_items
        WHERE buy_at IS NOT NULL AND buy_at >= NOW() - INTERVAL '90 days'
        GROUP BY DATE(buy_at)
        ORDER BY day
    """)
    by_day_buy = [dict(r) for r in cur.fetchall()]

    # ── 4. По дням: продажи ──────────────────────────────────────────────────
    cur.execute(f"""
        SELECT
            DATE(sell_at)                             AS day,
            COUNT(*)                                  AS sold_count,
            COALESCE(SUM(sell_price), 0)              AS revenue,
            COALESCE(SUM(buy_price), 0)               AS cogs,
            COALESCE(SUM(sell_price - buy_price), 0)  AS profit
        FROM {SCHEMA}.slshop_items
        WHERE sell_at IS NOT NULL AND sell_at >= NOW() - INTERVAL '90 days'
        GROUP BY DATE(sell_at)
        ORDER BY day
    """)
    by_day_sell = [dict(r) for r in cur.fetchall()]

    # ── 5. Баланс на начало и конец каждого месяца ───────────────────────────
    # «Стоимость склада на конец месяца M» = сумма buy_price товаров,
    #   которые были закуплены до конца M, но ещё не проданы к концу M
    cur.execute(f"""
        WITH months AS (
            SELECT DISTINCT DATE_TRUNC('month', created_at) AS m
            FROM {SCHEMA}.slshop_items
            WHERE created_at IS NOT NULL
            ORDER BY m
        )
        SELECT
            TO_CHAR(m, 'YYYY-MM')                       AS month,
            (SELECT COALESCE(SUM(i.buy_price), 0)
             FROM {SCHEMA}.slshop_items i
             WHERE i.created_at <= (m + INTERVAL '1 month' - INTERVAL '1 second')
               AND (i.sell_at IS NULL OR i.sell_at > (m + INTERVAL '1 month' - INTERVAL '1 second'))
               AND i.status != 'returned'
            )                                            AS stock_value_end,
            (SELECT COUNT(*)
             FROM {SCHEMA}.slshop_items i
             WHERE i.created_at <= (m + INTERVAL '1 month' - INTERVAL '1 second')
               AND (i.sell_at IS NULL OR i.sell_at > (m + INTERVAL '1 month' - INTERVAL '1 second'))
               AND i.status != 'returned'
            )                                            AS stock_count_end
        FROM months
        ORDER BY m
    """)
    monthly_balance = [dict(r) for r in cur.fetchall()]

    # ── 6. По категориям: текущие остатки ────────────────────────────────────
    cur.execute(f"""
        SELECT
            COALESCE(c.name, 'Без категории') AS category,
            COUNT(*)                          AS count,
            COALESCE(SUM(i.buy_price), 0)     AS buy_sum,
            COALESCE(SUM(i.sell_price), 0)    AS sell_sum,
            ROUND(AVG(i.buy_price))           AS avg_buy
        FROM {SCHEMA}.slshop_items i
        LEFT JOIN {SCHEMA}.slshop_categories c ON c.id = i.category_id
        WHERE i.status IN ('stock', 'showcase', 'consignment')
        GROUP BY c.name
        ORDER BY buy_sum DESC
    """)
    by_category = [dict(r) for r in cur.fetchall()]

    cur.close(); conn.close()

    # Вычисляем ROI: прибыль / вложения в проданный товар * 100
    cogs = float(totals.get('total_cogs') or 0)
    profit = float(totals.get('total_profit') or 0)
    roi = round(profit / cogs * 100, 1) if cogs > 0 else 0

    # Оборачиваемость: среднее кол-во дней от закупки до продажи (приближение)
    total_sold = int(totals.get('sold_count') or 0)
    total_stock = int(totals.get('in_stock_count') or 0)

    return _ok({
        'totals': {
            'started_at': totals['started_at'].isoformat() if totals.get('started_at') else None,
            'total_items': int(totals.get('total_items') or 0),
            'in_stock_count': total_stock,
            'sold_count': total_sold,
            'returned_count': int(totals.get('returned_count') or 0),
            'total_invested': float(totals.get('total_invested') or 0),
            'stock_value_buy': float(totals.get('stock_value_buy') or 0),
            'stock_value_sell': float(totals.get('stock_value_sell') or 0),
            'total_revenue': float(totals.get('total_revenue') or 0),
            'total_cogs': float(totals.get('total_cogs') or 0),
            'total_profit': float(totals.get('total_profit') or 0),
            'roi': roi,
        },
        'by_month_buy': by_month_buy,
        'by_month_sell': by_month_sell,
        'by_day_buy': by_day_buy,
        'by_day_sell': by_day_sell,
        'monthly_balance': monthly_balance,
        'by_category': by_category,
    })


def analytics_full(params):
    """Расширенная аналитика: периоды, разбивка по сотрудникам, филиалам, категориям, дням"""
    period = (params.get('period') or '30d').strip()
    now = datetime.utcnow()
    if period == 'today':
        date_from = now.strftime('%Y-%m-%d')
    elif period == 'yesterday':
        date_from = (now - timedelta(days=1)).strftime('%Y-%m-%d')
    elif period == '7d':
        date_from = (now - timedelta(days=7)).strftime('%Y-%m-%d')
    elif period == '30d':
        date_from = (now - timedelta(days=30)).strftime('%Y-%m-%d')
    elif period == 'year':
        date_from = (now - timedelta(days=365)).strftime('%Y-%m-%d')
    else:
        date_from = '2000-01-01'
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cond_op = f" WHERE o.created_at >= {_esc(date_from)}"
    # По сотрудникам
    cur.execute(
        f"SELECT COALESCE(o.employee_name,'—') AS employee, "
        f"COUNT(*) FILTER (WHERE o.op_type='sell') AS sold_count, "
        f"COALESCE(SUM(o.amount) FILTER (WHERE o.op_type='sell'),0) AS sold_sum, "
        f"COUNT(*) FILTER (WHERE o.op_type='buy') AS bought_count, "
        f"COALESCE(SUM(o.amount) FILTER (WHERE o.op_type='buy'),0) AS bought_sum "
        f"FROM {SCHEMA}.slshop_operations o {cond_op} GROUP BY o.employee_name ORDER BY sold_sum DESC"
    )
    by_employee = [dict(r) for r in cur.fetchall()]
    # По филиалам
    cur.execute(
        f"SELECT b.name AS branch, "
        f"COUNT(*) FILTER (WHERE o.op_type='sell') AS sold_count, "
        f"COALESCE(SUM(o.amount) FILTER (WHERE o.op_type='sell'),0) AS sold_sum, "
        f"COALESCE(SUM(o.amount) FILTER (WHERE o.op_type='buy'),0) AS bought_sum "
        f"FROM {SCHEMA}.slshop_operations o "
        f"LEFT JOIN {SCHEMA}.slshop_items i ON i.id=o.item_id "
        f"LEFT JOIN {SCHEMA}.slshop_branches b ON b.id=i.branch_id "
        f"{cond_op} GROUP BY b.name ORDER BY sold_sum DESC"
    )
    by_branch = [dict(r) for r in cur.fetchall()]
    # По дням
    cur.execute(
        f"SELECT DATE(o.created_at) AS d, "
        f"COUNT(*) FILTER (WHERE o.op_type='sell') AS sold_count, "
        f"COALESCE(SUM(o.amount) FILTER (WHERE o.op_type='sell'),0) AS sold_sum, "
        f"COALESCE(SUM(o.amount) FILTER (WHERE o.op_type='buy'),0) AS bought_sum "
        f"FROM {SCHEMA}.slshop_operations o {cond_op} GROUP BY 1 ORDER BY 1"
    )
    by_day = [dict(r) for r in cur.fetchall()]
    # Категории
    cur.execute(
        f"SELECT c.name AS category, COUNT(*) FILTER (WHERE o.op_type='sell') AS sold_count, "
        f"COALESCE(SUM(o.amount) FILTER (WHERE o.op_type='sell'),0) AS sold_sum "
        f"FROM {SCHEMA}.slshop_operations o "
        f"LEFT JOIN {SCHEMA}.slshop_items i ON i.id=o.item_id "
        f"LEFT JOIN {SCHEMA}.slshop_categories c ON c.id=i.category_id "
        f"{cond_op} GROUP BY c.name ORDER BY sold_sum DESC LIMIT 20"
    )
    by_category = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return _ok({
        'period': period, 'date_from': date_from,
        'by_employee': by_employee,
        'by_branch': by_branch,
        'by_day': by_day,
        'by_category': by_category,
    })


def accounting_summary(params):
    """Бухгалтерия: оборот, прибыль, налоговая база за период по филиалам и кассам."""
    period = (params.get('period') or '30d').strip()
    now = datetime.utcnow()
    if period == 'custom':
        df_raw = (params.get('date_from') or '').strip()
        try:
            datetime.strptime(df_raw[:10], '%Y-%m-%d')
            date_from = df_raw[:10]
        except Exception:
            date_from = (now - timedelta(days=30)).strftime('%Y-%m-%d')
    elif period == 'today':
        date_from = now.strftime('%Y-%m-%d')
    elif period == 'yesterday':
        date_from = (now - timedelta(days=1)).strftime('%Y-%m-%d')
    elif period == '7d':
        date_from = (now - timedelta(days=7)).strftime('%Y-%m-%d')
    elif period == '30d':
        date_from = (now - timedelta(days=30)).strftime('%Y-%m-%d')
    elif period == 'year':
        date_from = (now - timedelta(days=365)).strftime('%Y-%m-%d')
    elif period == 'all':
        date_from = '2000-01-01'
    else:
        date_from = (now - timedelta(days=30)).strftime('%Y-%m-%d')
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    # 1) Выручка от продаж и общий объём закупок товаров за период
    cur.execute(
        f"SELECT "
        f"COALESCE(SUM(amount) FILTER (WHERE op_type='sell'),0) AS revenue, "
        f"COALESCE(SUM(amount) FILTER (WHERE op_type='buy'),0) AS purchases, "
        f"COUNT(*) FILTER (WHERE op_type='sell') AS sales_count, "
        f"COUNT(*) FILTER (WHERE op_type='buy') AS buys_count "
        f"FROM {SCHEMA}.slshop_operations WHERE created_at >= {_esc(date_from)}"
    )
    totals = dict(cur.fetchone() or {})
    # 2) Себестоимость проданных за период (COGS): для каждой 'sell'-операции
    # берём buy_price связанного товара. Это отдельно от закупочного оборота.
    cur.execute(
        f"SELECT COALESCE(SUM(COALESCE(i.buy_price, 0)), 0) AS cogs "
        f"FROM {SCHEMA}.slshop_operations o "
        f"LEFT JOIN {SCHEMA}.slshop_items i ON i.id = o.item_id "
        f"WHERE o.op_type='sell' AND o.created_at >= {_esc(date_from)}"
    )
    cogs_row = cur.fetchone() or {}
    cogs = float(cogs_row.get('cogs') or 0)
    # 3) Операционные расходы (зарплаты, аренда и т.п. — кассовые out-движения)
    cur.execute(
        f"SELECT COALESCE(SUM(amount), 0) AS opex "
        f"FROM {SCHEMA}.slshop_cash_movements "
        f"WHERE direction='out' AND created_at >= {_esc(date_from)}"
    )
    opex_row = cur.fetchone() or {}
    opex = float(opex_row.get('opex') or 0)
    # 3.1) СмартЛомбард: детализация по закрытым договорам за период
    # Прибыль = сколько вернул клиент - сколько ему выдали (тело)
    contract_closed_count = 0
    contract_issued = 0.0
    contract_returned = 0.0
    contract_profit = 0.0
    contract_active_count = 0
    contract_active_issued = 0.0
    try:
        cur.execute(
            f"SELECT COUNT(*) AS cnt, "
            f"COALESCE(SUM(amount),0) AS issued, "
            f"COALESCE(SUM(paid_total),0) AS returned "
            f"FROM {SCHEMA}.contracts_14d "
            f"WHERE status='closed' AND closed_at>={_esc(date_from)}"
        )
        cr = cur.fetchone() or {}
        contract_closed_count = int(cr.get('cnt') or 0)
        contract_issued = float(cr.get('issued') or 0)
        contract_returned = float(cr.get('returned') or 0)
        contract_profit = contract_returned - contract_issued
        cur.execute(
            f"SELECT COUNT(*) AS cnt, COALESCE(SUM(amount),0) AS issued "
            f"FROM {SCHEMA}.contracts_14d WHERE status='active'"
        )
        cr2 = cur.fetchone() or {}
        contract_active_count = int(cr2.get('cnt') or 0)
        contract_active_issued = float(cr2.get('issued') or 0)
    except Exception:
        pass
    contract_income = contract_profit
    # 4) Касса по филиалам (текущий баланс + движение за период)
    cur.execute(
        f"SELECT b.name AS branch, a.name AS account, a.balance, "
        f"COALESCE((SELECT SUM(amount) FROM {SCHEMA}.slshop_cash_movements m WHERE m.account_id=a.id AND m.direction='in' AND m.created_at >= {_esc(date_from)}),0) AS in_sum, "
        f"COALESCE((SELECT SUM(amount) FROM {SCHEMA}.slshop_cash_movements m WHERE m.account_id=a.id AND m.direction='out' AND m.created_at >= {_esc(date_from)}),0) AS out_sum "
        f"FROM {SCHEMA}.slshop_cash_accounts a LEFT JOIN {SCHEMA}.slshop_branches b ON b.id=a.branch_id "
        f"WHERE a.is_active=true ORDER BY b.name"
    )
    cash = [dict(r) for r in cur.fetchall()]
    # 5) Операционные расходы по категориям (из кассы)
    cur.execute(
        f"SELECT category, COUNT(*) AS cnt, COALESCE(SUM(amount),0) AS sum "
        f"FROM {SCHEMA}.slshop_cash_movements "
        f"WHERE direction='out' AND created_at >= {_esc(date_from)} "
        f"GROUP BY category ORDER BY sum DESC"
    )
    expenses = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    revenue = float(totals.get('revenue') or 0)
    purchases = float(totals.get('purchases') or 0)
    # Прибыль б/у — выручка с продаж минус себестоимость проданного
    profit_buy_used = revenue - cogs
    # Общая маржа = прибыль б/у + прибыль СмартЛомбарда
    gross_profit = profit_buy_used + contract_profit
    # Чистая прибыль с учётом расходов кассы
    net_profit = gross_profit - opex
    # Полные затраты периода (для совместимости со старым полем spent)
    spent = purchases + opex
    return _ok({
        'period': period, 'date_from': date_from,
        'revenue': revenue,
        'purchases': purchases,         # сколько потратили на закупку товаров
        'cogs': cogs,                   # себестоимость проданных за период
        'opex': opex,                   # операционные расходы (касса out)
        'profit_used': profit_buy_used, # прибыль чисто от продаж б/у
        # СмартЛомбард — детализация
        'contract_closed_count': contract_closed_count,
        'contract_issued': contract_issued,
        'contract_returned': contract_returned,
        'contract_profit': contract_profit,
        'contract_active_count': contract_active_count,
        'contract_active_issued': contract_active_issued,
        'contract_income': contract_income,        # = contract_profit, для совместимости
        'gross_profit': gross_profit,   # б/у + ломбард
        'profit': net_profit,           # gross - opex
        'spent': spent,                 # обратная совместимость
        'sales_count': int(totals.get('sales_count') or 0),
        'buys_count': int(totals.get('buys_count') or 0),
        'cash_by_branch': cash,
        'expenses_by_category': expenses,
    })


def list_favorites(employee):
    if not employee:
        return []
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        f"SELECT * FROM {SCHEMA}.slshop_favorites WHERE employee_id={int(employee['id'])} "
        f"ORDER BY sort_order, id"
    )
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [dict(r) for r in rows]


def favorite_add(body, employee):
    if not employee:
        return _err(401, 'Не авторизован')
    label = (body.get('label') or '').strip()
    kind = body.get('kind') or 'page'
    if not label:
        return _err(400, 'label обязателен')
    conn = get_conn(); cur = conn.cursor()
    cur.execute(
        f"INSERT INTO {SCHEMA}.slshop_favorites (employee_id, kind, ref_id, label, url, icon) "
        f"VALUES ({int(employee['id'])}, {_esc(kind)}, {_esc(body.get('ref_id'))}, {_esc(label)}, "
        f"{_esc(body.get('url'))}, {_esc(body.get('icon') or 'Star')}) RETURNING id"
    )
    nid = cur.fetchone()[0]
    conn.commit(); cur.close(); conn.close()
    return _ok({'id': nid})


def favorite_remove(body, employee):
    if not employee:
        return _err(401, 'Не авторизован')
    fid = body.get('id')
    if not fid:
        return _err(400, 'id обязателен')
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"DELETE FROM {SCHEMA}.slshop_favorites WHERE id={int(fid)} AND employee_id={int(employee['id'])}")
    conn.commit(); cur.close(); conn.close()
    return _ok({'ok': True})


def operation_delete(body, employee):
    if not employee or employee.get('role') != 'owner':
        return _err(403, 'Удалять операции может только владелец')
    op_id = body.get('id')
    if not op_id:
        return _err(400, 'id обязателен')
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT op_type, item_id, status_from, amount FROM {SCHEMA}.slshop_operations WHERE id={int(op_id)}")
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close(); return _err(404, 'Операция не найдена')
    op_type, item_id, status_from, amount = row[0], row[1], row[2], row[3]
    op_id_int = int(op_id)

    # 1) Откатываем эффект операции на товар (статус возвращаем)
    if op_type == 'sell' and item_id:
        prev = status_from or 'stock'
        cur.execute(
            f"UPDATE {SCHEMA}.slshop_items SET status={_esc(prev)}, sell_at=NULL, sell_operation_id=NULL, updated_at=NOW() WHERE id={int(item_id)}"
        )
    elif op_type == 'return' and item_id:
        prev = status_from or 'stock'
        cur.execute(
            f"UPDATE {SCHEMA}.slshop_items SET status={_esc(prev)}, updated_at=NOW() WHERE id={int(item_id)}"
        )
    elif op_type == 'move' and item_id and status_from:
        cur.execute(
            f"UPDATE {SCHEMA}.slshop_items SET status={_esc(status_from)}, updated_at=NOW() WHERE id={int(item_id)}"
        )
    elif op_type == 'buy' and item_id:
        # buy → у товара затёрта buy_operation_id, item остаётся
        cur.execute(
            f"UPDATE {SCHEMA}.slshop_items SET buy_operation_id=NULL WHERE id={int(item_id)} AND buy_operation_id={op_id_int}"
        )

    # 2) Обнуляем все возможные FK-ссылки на эту операцию, чтобы DELETE не упал
    cur.execute(f"UPDATE {SCHEMA}.slshop_items SET sell_operation_id=NULL WHERE sell_operation_id={op_id_int}")
    cur.execute(f"UPDATE {SCHEMA}.slshop_items SET buy_operation_id=NULL WHERE buy_operation_id={op_id_int}")
    cur.execute(f"UPDATE {SCHEMA}.slshop_cash_movements SET related_op_id=NULL WHERE related_op_id={op_id_int}")
    cur.execute(f"UPDATE {SCHEMA}.slshop_operations SET related_op_id=NULL WHERE related_op_id={op_id_int}")

    # 3) Откатим связанное движение кассы (если оно было автоматическим от продажи/скупки)
    cur.execute(
        f"SELECT id, account_id, direction, amount FROM {SCHEMA}.slshop_cash_movements "
        f"WHERE related_item_id IS NOT NULL AND is_auto = TRUE AND related_op_id IS NULL "
        f"AND created_at >= NOW() - INTERVAL '5 minutes' LIMIT 0"
    )
    # Удалять движения кассы автоматически не будем — оставим решение за пользователем

    # 4) Удаляем саму операцию
    cur.execute(f"DELETE FROM {SCHEMA}.slshop_operations WHERE id={op_id_int}")

    # 5) Лог
    _log_event(cur, 'remove', 'operation', op_id_int, f'Удалена операция #{op_id_int}',
               f'Тип: {op_type}, сумма: {amount}', float(amount or 0), None, employee)

    conn.commit(); cur.close(); conn.close()
    return _ok({'ok': True})


def list_sold(params):
    period = (params.get('period') or '30d').strip()
    now = datetime.utcnow()
    if period == 'today':
        date_from = now.strftime('%Y-%m-%d')
    elif period == 'yesterday':
        date_from = (now - timedelta(days=1)).strftime('%Y-%m-%d')
    elif period == '7d':
        date_from = (now - timedelta(days=7)).strftime('%Y-%m-%d')
    elif period == '30d':
        date_from = (now - timedelta(days=30)).strftime('%Y-%m-%d')
    elif period == 'year':
        date_from = (now - timedelta(days=365)).strftime('%Y-%m-%d')
    else:
        date_from = '2000-01-01'
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        f"SELECT i.id, i.title, i.specs_short, i.imei, i.sku, i.sell_price, i.buy_price, i.sell_at, "
        f"i.category_id, c.name AS category_name, c.path AS category_path, "
        f"o.id AS operation_id, o.amount, o.payment_method, o.contract_number, o.employee_name, "
        f"o.client_id, cl.full_name AS client_name, cl.phone AS client_phone, "
        f"b.name AS branch_name, b.address AS branch_address "
        f"FROM {SCHEMA}.slshop_items i "
        f"JOIN {SCHEMA}.slshop_operations o ON o.id=i.sell_operation_id "
        f"LEFT JOIN {SCHEMA}.slshop_categories c ON c.id=i.category_id "
        f"LEFT JOIN {SCHEMA}.slshop_clients cl ON cl.id=o.client_id "
        f"LEFT JOIN {SCHEMA}.slshop_branches b ON b.id=i.branch_id "
        f"WHERE i.status='sold' AND i.sell_at >= {_esc(date_from)} "
        f"ORDER BY i.sell_at DESC LIMIT 200"
    )
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [dict(r) for r in rows]


def list_bought(params):
    period = (params.get('period') or '30d').strip()
    now = datetime.utcnow()
    if period == 'today':
        date_from = now.strftime('%Y-%m-%d')
    elif period == 'yesterday':
        date_from = (now - timedelta(days=1)).strftime('%Y-%m-%d')
    elif period == '7d':
        date_from = (now - timedelta(days=7)).strftime('%Y-%m-%d')
    elif period == '30d':
        date_from = (now - timedelta(days=30)).strftime('%Y-%m-%d')
    elif period == 'year':
        date_from = (now - timedelta(days=365)).strftime('%Y-%m-%d')
    else:
        date_from = '2000-01-01'
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        f"SELECT i.id, i.title, i.specs_short, i.imei, i.sku, i.buy_price, i.sell_price, i.buy_at, i.status, "
        f"i.category_id, c.name AS category_name, c.path AS category_path, "
        f"o.id AS operation_id, o.amount, o.employee_name, "
        f"o.client_id, cl.full_name AS client_name, cl.phone AS client_phone, "
        f"b.name AS branch_name, b.address AS branch_address "
        f"FROM {SCHEMA}.slshop_items i "
        f"LEFT JOIN {SCHEMA}.slshop_operations o ON o.id=i.buy_operation_id "
        f"LEFT JOIN {SCHEMA}.slshop_categories c ON c.id=i.category_id "
        f"LEFT JOIN {SCHEMA}.slshop_clients cl ON cl.id=o.client_id "
        f"LEFT JOIN {SCHEMA}.slshop_branches b ON b.id=i.branch_id "
        f"WHERE i.buy_at IS NOT NULL AND i.buy_at >= {_esc(date_from)} "
        f"ORDER BY i.buy_at DESC LIMIT 200"
    )
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [dict(r) for r in rows]


def revision_finish(body):
    rev_id = body.get('id')
    if not rev_id:
        return _err(400, 'id обязателен')
    conn = get_conn(); cur = conn.cursor()
    cur.execute(
        f"UPDATE {SCHEMA}.slshop_revision_items SET state='missing' "
        f"WHERE revision_id={int(rev_id)} AND state='pending'"
    )
    cur.execute(
        f"UPDATE {SCHEMA}.slshop_revisions SET "
        f"total_found=(SELECT COUNT(*) FROM {SCHEMA}.slshop_revision_items WHERE revision_id={int(rev_id)} AND state='found'), "
        f"total_missing=(SELECT COUNT(*) FROM {SCHEMA}.slshop_revision_items WHERE revision_id={int(rev_id)} AND state='missing'), "
        f"total_extra=(SELECT COUNT(*) FROM {SCHEMA}.slshop_revision_items WHERE revision_id={int(rev_id)} AND state='extra'), "
        f"status='closed', finished_at=NOW() WHERE id={int(rev_id)}"
    )
    conn.commit(); cur.close(); conn.close()
    return _ok({'ok': True})


# ============ Clients ============
def list_clients(params):
    q = (params.get('q') or '').strip()
    cond = ''
    if q:
        like = f"%{q}%"
        cond = f" WHERE full_name ILIKE {_esc(like)} OR phone ILIKE {_esc(like)}"
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(f"SELECT * FROM {SCHEMA}.slshop_clients{cond} ORDER BY created_at DESC LIMIT 200")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [dict(r) for r in rows]


def upsert_client(body):
    cid = body.get('id')
    full_name = (body.get('full_name') or '').strip()
    phone = (body.get('phone') or '').strip()
    if not full_name:
        return _err(400, 'ФИО обязательно')
    fields = {
        'full_name': full_name,
        'phone': phone,
        'passport_series': body.get('passport_series'),
        'passport_number': body.get('passport_number'),
        'passport_issued_by': body.get('passport_issued_by'),
        'passport_issued_date': body.get('passport_issued_date') or None,
        'address': body.get('address'),
        'birth_date': body.get('birth_date') or None,
        'notes': body.get('notes'),
    }
    # Фото паспорта (опциональные поля). Не пишем None, чтобы не затирать
    # существующие фото при обновлении только текстовых полей.
    for fkey in ('passport_photo_url', 'passport_photo2_url', 'face_photo_url'):
        if fkey in body and body.get(fkey):
            fields[fkey] = body.get(fkey)
    conn = get_conn(); cur = conn.cursor()
    if cid:
        sets = ', '.join([f"{k}={_esc(v)}" for k, v in fields.items()])
        cur.execute(f"UPDATE {SCHEMA}.slshop_clients SET {sets}, updated_at=NOW() WHERE id={int(cid)} RETURNING id")
        row = cur.fetchone()
        conn.commit(); cur.close(); conn.close()
        return _ok({'id': row[0] if row else cid})
    else:
        cols = ', '.join(fields.keys())
        vals = ', '.join([_esc(v) for v in fields.values()])
        cur.execute(f"INSERT INTO {SCHEMA}.slshop_clients ({cols}) VALUES ({vals}) RETURNING id")
        nid = cur.fetchone()[0]
        conn.commit(); cur.close(); conn.close()
        return _ok({'id': nid})


# ============ Client passport photo upload + AI OCR ============
PASSPORT_OCR_PROMPT = (
    "Ты — помощник по распознаванию российского паспорта. На фото — страница с фотографией владельца "
    "(2-3 разворот) или страница регистрации. Извлеки текстовые данные и верни СТРОГО JSON без комментариев.\n\n"
    "Поля:\n"
    "- full_name: ФИО полностью одной строкой в формате 'Фамилия Имя Отчество' (с заглавных букв)\n"
    "- series: серия паспорта (4 цифры, можно с пробелом, например '12 34')\n"
    "- number: номер паспорта (6 цифр)\n"
    "- issued_by: кем выдан, как написано в паспорте, одной строкой\n"
    "- issued_date: дата выдачи в формате YYYY-MM-DD (если видна)\n"
    "- birth_date: дата рождения в формате YYYY-MM-DD (если видна)\n"
    "- address: адрес регистрации одной строкой (если виден на странице с пропиской)\n\n"
    "Если поле не видно на фото — верни пустую строку для строк или null для дат. "
    "Не выдумывай данные, верни только то, что реально читается на фото."
)


def _get_s3():
    import boto3
    return boto3.client(
        's3',
        endpoint_url='https://bucket.poehali.dev',
        aws_access_key_id=os.environ['AWS_ACCESS_KEY_ID'],
        aws_secret_access_key=os.environ['AWS_SECRET_ACCESS_KEY'],
    )


def _cdn_url(key: str) -> str:
    return f"https://cdn.poehali.dev/projects/{os.environ['AWS_ACCESS_KEY_ID']}/bucket/{key}"


def _ocr_passport_via_ai(data_url: str) -> dict:
    """Распознаёт паспорт через GPT-4o (Polza.ai). Возвращает dict с полями (можно пустыми).
    Если ключа нет или распознавание упало — возвращает пустой словарь, не падает."""
    api_key = os.environ.get('POLZA_AI_API_KEY', '')
    if not api_key:
        return {}
    try:
        payload = json.dumps({
            "model": "gpt-4o-mini",
            "messages": [{
                "role": "user",
                "content": [
                    {"type": "text", "text": PASSPORT_OCR_PROMPT},
                    {"type": "image_url", "image_url": {"url": data_url}},
                ],
            }],
            "max_tokens": 600,
            "temperature": 0.1,
            "response_format": {"type": "json_object"},
        }).encode('utf-8')
        req = urllib.request.Request(
            'https://api.polza.ai/v1/chat/completions',
            data=payload,
            headers={
                'Content-Type': 'application/json',
                'Authorization': f'Bearer {api_key}',
            },
        )
        with urllib.request.urlopen(req, timeout=40) as resp:
            data = json.loads(resp.read())
        raw = (data.get('choices') or [{}])[0].get('message', {}).get('content', '').strip()
        if raw.startswith('```'):
            raw = raw.strip('`')
            if raw.lower().startswith('json'):
                raw = raw[4:]
            raw = raw.strip()
        parsed = json.loads(raw)
        # Нормализуем ключи: серия без пробелов будет лучше для отображения, но оставим как есть
        result = {
            'full_name': (parsed.get('full_name') or '').strip(),
            'series': (parsed.get('series') or '').strip().replace(' ', ''),
            'number': (parsed.get('number') or '').strip().replace(' ', ''),
            'issued_by': (parsed.get('issued_by') or '').strip(),
            'issued_date': parsed.get('issued_date') or None,
            'birth_date': parsed.get('birth_date') or None,
            'address': (parsed.get('address') or '').strip(),
        }
        return result
    except Exception as e:
        # AI-сбой не должен ломать загрузку самого фото
        return {'_ocr_error': str(e)[:200]}


def client_passport_upload(body, employee):
    """Принимает фото паспорта в base64, грузит в S3 и (если есть AI-ключ)
    распознаёт паспортные данные. Возвращает {url, passport_data}.

    body:
      - image_base64: data:image/...;base64,... или чистый base64
      - field: 'passport_photo_url' | 'passport_photo2_url' | 'face_photo_url' (опц., для S3-пути)
      - recognize: bool (по умолчанию true) — запускать ли OCR
    """
    image_b64 = body.get('image_base64') or ''
    if not image_b64:
        return _err(400, 'image_base64 обязателен')
    field = (body.get('field') or 'passport_photo_url').strip()
    recognize = body.get('recognize')
    if recognize is None:
        recognize = True

    # Подготавливаем data_url для AI и raw bytes для S3
    if image_b64.startswith('data:'):
        data_url = image_b64
        # вырезаем base64-часть
        try:
            _, b64_part = image_b64.split(',', 1)
        except ValueError:
            return _err(400, 'некорректный формат image_base64')
    else:
        data_url = 'data:image/jpeg;base64,' + image_b64
        b64_part = image_b64

    try:
        raw_bytes = base64.b64decode(b64_part)
    except Exception:
        return _err(400, 'не удалось декодировать base64')
    if not raw_bytes or len(raw_bytes) < 200:
        return _err(400, 'фото слишком маленькое')

    # Загрузка в S3
    suffix = 'jpg'
    if 'image/png' in data_url[:40]:
        suffix = 'png'
    elif 'image/webp' in data_url[:40]:
        suffix = 'webp'
    today = datetime.utcnow().strftime('%Y%m%d')
    key = f"passport/{today}/{field}_{uuid.uuid4().hex[:12]}.{suffix}"
    try:
        s3 = _get_s3()
        s3.put_object(
            Bucket='files',
            Key=key,
            Body=raw_bytes,
            ContentType=f'image/{suffix}',
        )
    except Exception as e:
        return _err(500, f'Не удалось сохранить фото: {str(e)[:200]}')
    url = _cdn_url(key)

    # Распознавание (опционально, на странице 2-3 паспорта)
    passport_data = {}
    if recognize and field in ('passport_photo_url', 'passport_photo2_url'):
        passport_data = _ocr_passport_via_ai(data_url)

    # Поиск дубля клиента по распознанным данным.
    # Приоритет: 1) серия+номер (точно), 2) ФИО+ДР, 3) только ФИО.
    matches = []
    if passport_data and not passport_data.get('_ocr_error'):
        try:
            conn = get_conn()
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            series = (passport_data.get('series') or '').strip().replace(' ', '')
            number = (passport_data.get('number') or '').strip().replace(' ', '')
            full_name = (passport_data.get('full_name') or '').strip()
            birth_date = passport_data.get('birth_date')

            base_select = (
                f"SELECT id, full_name, phone, passport_series, passport_number, "
                f"passport_issued_by, passport_issued_date, address, birth_date, "
                f"passport_photo_url, passport_photo2_url, face_photo_url "
                f"FROM {SCHEMA}.slshop_clients "
            )
            # Точный матч серия+номер
            if series and number:
                cur.execute(
                    base_select +
                    f"WHERE REPLACE(COALESCE(passport_series,''),' ','')={_esc(series)} "
                    f"AND REPLACE(COALESCE(passport_number,''),' ','')={_esc(number)} "
                    f"ORDER BY updated_at DESC LIMIT 5"
                )
                matches = [dict(r) for r in cur.fetchall()]
            # ФИО + ДР
            if not matches and full_name and birth_date:
                cur.execute(
                    base_select +
                    f"WHERE LOWER(full_name)=LOWER({_esc(full_name)}) "
                    f"AND birth_date={_esc(birth_date)} LIMIT 5"
                )
                matches = [dict(r) for r in cur.fetchall()]
            # Только ФИО (несколько однофамильцев)
            if not matches and full_name:
                cur.execute(
                    base_select +
                    f"WHERE LOWER(full_name)=LOWER({_esc(full_name)}) "
                    f"ORDER BY updated_at DESC LIMIT 5"
                )
                matches = [dict(r) for r in cur.fetchall()]

            # Подтянем краткую статистику по каждому: сколько у нас от него
            # принято товаров, на какую сумму, когда последний приход.
            if matches:
                ids = ', '.join(str(int(m['id'])) for m in matches)
                cur.execute(
                    f"SELECT buy_client_id AS cid, "
                    f"COUNT(*) AS items_count, "
                    f"COALESCE(SUM(buy_price), 0) AS total_buy, "
                    f"COALESCE(SUM(sell_price), 0) AS total_sell, "
                    f"MAX(buy_at) AS last_buy_at "
                    f"FROM {SCHEMA}.slshop_items "
                    f"WHERE buy_client_id IN ({ids}) "
                    f"GROUP BY buy_client_id"
                )
                stats_by_id = {int(r['cid']): r for r in cur.fetchall()}
                for m in matches:
                    s = stats_by_id.get(int(m['id']))
                    if s:
                        m['stats'] = {
                            'items_count': int(s['items_count'] or 0),
                            'total_buy': float(s['total_buy'] or 0),
                            'total_sell': float(s['total_sell'] or 0),
                            'last_buy_at': s['last_buy_at'].isoformat() if s.get('last_buy_at') else None,
                        }
                    else:
                        m['stats'] = {'items_count': 0, 'total_buy': 0, 'total_sell': 0, 'last_buy_at': None}
            cur.close(); conn.close()
        except Exception:
            matches = matches or []

    return _ok({
        'url': url,
        'field': field,
        'passport_data': passport_data,
        'matches': matches,
    })


# ============ Накладные: экспорт XLSX и распознавание по фото ============
INVOICE_OCR_PROMPT = (
    "На фото — приёмная накладная (список товаров на скупку). "
    "Извлеки СТРОКИ таблицы и верни СТРОГО JSON-массив объектов в поле items. "
    "Поля каждого объекта:\n"
    "- title: наименование товара одной строкой (например 'Awei Y680 колонка портативная')\n"
    "- quantity: количество, целое число (если не указано — поставь 1)\n"
    "- buy_price: цена закупки за 1 шт, целое число в рублях (без валюты)\n"
    "- color: цвет, если указан (иначе пустая строка)\n\n"
    "Не выдумывай. Если каких-то полей не видно — оставь пустыми/0. "
    "Не добавляй комментарии. Верни строго JSON: {\"items\":[{...}, ...]}."
)


def _invoice_ocr_via_ai(data_url: str) -> list:
    """Распознаёт строки накладной через GPT-4o (Polza.ai). Если ключа нет
    или произошёл сбой — возвращает пустой список."""
    api_key = os.environ.get('POLZA_AI_API_KEY', '')
    if not api_key:
        return []
    try:
        payload = json.dumps({
            "model": "gpt-4o-mini",
            "messages": [{
                "role": "user",
                "content": [
                    {"type": "text", "text": INVOICE_OCR_PROMPT},
                    {"type": "image_url", "image_url": {"url": data_url}},
                ],
            }],
            "max_tokens": 3000,
            "temperature": 0.1,
            "response_format": {"type": "json_object"},
        }).encode('utf-8')
        req = urllib.request.Request(
            'https://api.polza.ai/v1/chat/completions',
            data=payload,
            headers={
                'Content-Type': 'application/json',
                'Authorization': f'Bearer {api_key}',
            },
        )
        with urllib.request.urlopen(req, timeout=60) as resp:
            data = json.loads(resp.read())
        raw = (data.get('choices') or [{}])[0].get('message', {}).get('content', '').strip()
        if raw.startswith('```'):
            raw = raw.strip('`')
            if raw.lower().startswith('json'):
                raw = raw[4:]
            raw = raw.strip()
        parsed = json.loads(raw)
        items = parsed.get('items') or []
        result = []
        for it in items:
            try:
                title = str(it.get('title') or '').strip()
                qty = int(it.get('quantity') or 1)
                buy = float(str(it.get('buy_price') or '0').replace(',', '.').replace(' ', '') or 0)
                if not title:
                    continue
                if qty < 1:
                    qty = 1
                result.append({
                    'title': title,
                    'quantity': qty,
                    'buy_price': int(buy),
                    'color': str(it.get('color') or '').strip(),
                })
            except (TypeError, ValueError):
                continue
        return result
    except Exception:
        return []


def _auto_sell_price(buy_price: float) -> int:
    """Правило авто-розницы: ≤500 ₽ → ×3, иначе ×2."""
    b = float(buy_price or 0)
    if b <= 0:
        return 0
    return int(b * 3 if b <= 500 else b * 2)


def items_invoice_recognize(body, employee):
    """Принимает фото накладной (image_base64), грузит в S3, прогоняет через
    GPT-4o Vision и возвращает массив распознанных строк (без сохранения).
    Сотрудник проверит и подтвердит создание через items_bulk_create."""
    image_b64 = body.get('image_base64') or ''
    if not image_b64:
        return _err(400, 'image_base64 обязателен')
    if image_b64.startswith('data:'):
        data_url = image_b64
        try:
            _, b64_part = image_b64.split(',', 1)
        except ValueError:
            return _err(400, 'некорректный формат image_base64')
    else:
        data_url = 'data:image/jpeg;base64,' + image_b64
        b64_part = image_b64

    try:
        raw_bytes = base64.b64decode(b64_part)
    except Exception:
        return _err(400, 'не удалось декодировать base64')
    if not raw_bytes or len(raw_bytes) < 500:
        return _err(400, 'фото слишком маленькое')

    # Сохраняем фото в S3 для истории
    today = datetime.utcnow().strftime('%Y%m%d')
    key = f"invoices/{today}/invoice_{uuid.uuid4().hex[:12]}.jpg"
    photo_url = None
    try:
        s3 = _get_s3()
        s3.put_object(Bucket='files', Key=key, Body=raw_bytes, ContentType='image/jpeg')
        photo_url = _cdn_url(key)
    except Exception:
        photo_url = None

    items = _invoice_ocr_via_ai(data_url)
    # Добавим рекомендованную розницу к каждой строке
    for it in items:
        it['suggested_sell_price'] = _auto_sell_price(it.get('buy_price') or 0)

    return _ok({
        'photo_url': photo_url,
        'items': items,
        'count': len(items),
    })


def items_bulk_create(body, employee):
    """Массовое создание товаров из распознанной накладной.
    body: { items: [{title, quantity, buy_price, sell_price?, color?, brand?}, ...],
            branch_id, status, source }
    Возвращает { created: [...], errors: [...], total_count, total_qty, total_buy }"""
    items_in = body.get('items') or []
    if not isinstance(items_in, list) or not items_in:
        return _err(400, 'items должен быть непустым массивом')
    branch_id = body.get('branch_id')
    status = (body.get('status') or 'stock').strip()
    source = (body.get('source') or 'buyout').strip()
    default_brand = (body.get('brand') or '').strip()
    default_category_id = body.get('category_id')

    created = []
    errors = []
    total_qty = 0
    total_buy = 0.0

    for idx, it in enumerate(items_in):
        title = str(it.get('title') or '').strip()
        if not title:
            errors.append({'index': idx, 'error': 'пустое наименование'})
            continue
        try:
            qty = int(it.get('quantity') or 1)
        except (TypeError, ValueError):
            qty = 1
        if qty < 1:
            qty = 1
        try:
            buy = float(str(it.get('buy_price') or 0).replace(',', '.').replace(' ', '') or 0)
        except (TypeError, ValueError):
            buy = 0.0
        # Розница: если передана — используем, иначе авто-расчёт
        sell_raw = it.get('sell_price')
        if sell_raw in (None, '', 0, '0'):
            sell = _auto_sell_price(buy)
        else:
            try:
                sell = int(float(str(sell_raw).replace(',', '.').replace(' ', '')))
            except (TypeError, ValueError):
                sell = _auto_sell_price(buy)

        # Используем уже существующий create_item для единообразия
        sub_body = {
            'title': title,
            'brand': str(it.get('brand') or default_brand or '').strip(),
            'category_id': it.get('category_id') or default_category_id or None,
            'color': str(it.get('color') or '').strip(),
            'buy_price': buy,
            'sell_price': sell,
            'quantity': qty,
            'status': status,
            'source': source,
            'branch_id': branch_id,
        }
        res = create_item(sub_body, employee)
        try:
            data = json.loads(res.get('body', '{}'))
        except Exception:
            data = {}
        if res.get('statusCode') == 200 and data.get('id'):
            created.append({
                'index': idx,
                'id': data['id'],
                'sku': data.get('sku'),
                'title': title,
                'quantity': qty,
                'buy_price': buy,
                'sell_price': sell,
            })
            total_qty += qty
            total_buy += buy * qty
        else:
            errors.append({'index': idx, 'title': title, 'error': data.get('error') or 'Не удалось создать'})

    return _ok({
        'created': created,
        'errors': errors,
        'total_count': len(created),
        'total_qty': total_qty,
        'total_buy': total_buy,
    })


def items_invoice_export(params):
    """Экспортирует список товаров в XLSX в формате приёмной накладной.
    Параметры: date_from, date_to (YYYY-MM-DD), brand, branch_id, status.
    Возвращает base64 + filename."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except Exception:
        return _err(500, 'openpyxl не установлен')

    date_from = (params.get('date_from') or '').strip()
    date_to = (params.get('date_to') or '').strip()
    brand = (params.get('brand') or '').strip()
    branch_id = params.get('branch_id')
    status = (params.get('status') or '').strip()

    where = []
    if date_from:
        where.append(f"i.created_at::date >= {_esc(date_from)}::date")
    if date_to:
        where.append(f"i.created_at::date <= {_esc(date_to)}::date")
    if brand:
        where.append(f"LOWER(COALESCE(i.brand,''))=LOWER({_esc(brand)})")
    if branch_id:
        try:
            where.append(f"i.branch_id={int(branch_id)}")
        except (TypeError, ValueError):
            pass
    if status:
        where.append(f"i.status={_esc(status)}")
    wsql = (' WHERE ' + ' AND '.join(where)) if where else ''

    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        f"SELECT i.id, i.sku, i.title, i.brand, i.color, i.buy_price, i.sell_price, "
        f"COALESCE(i.quantity, 1) AS quantity, i.created_at, "
        f"b.name AS branch_name "
        f"FROM {SCHEMA}.slshop_items i "
        f"LEFT JOIN {SCHEMA}.slshop_branches b ON b.id=i.branch_id "
        f"{wsql} ORDER BY i.created_at ASC, i.id ASC"
    )
    rows = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()

    # Формируем XLSX в формате накладной
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Накладная'

    bold = Font(bold=True, size=11)
    big = Font(bold=True, size=13)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_a = Alignment(horizontal='right', vertical='center')
    thin = Side(style='thin', color='000000')
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill('solid', fgColor='FFE699')

    # Шапка
    title_text = 'НАКЛАДНАЯ ПРИЁМА ТОВАРА'
    if brand:
        title_text += f' — {brand}'
    if date_from or date_to:
        title_text += f' ({date_from or "..."} — {date_to or "..."})'
    ws.merge_cells('A1:G1')
    ws['A1'] = title_text
    ws['A1'].font = big
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 24

    headers = ['№', 'Наименование товара', 'Цвет', 'Кол-во', 'Цена ₽', 'Сумма ₽', 'SKU']
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col_idx, value=h)
        c.font = bold
        c.alignment = center
        c.border = box
        c.fill = head_fill

    # Ширина колонок
    widths = [5, 50, 14, 9, 12, 14, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    total_qty = 0
    total_buy = 0.0
    total_sell = 0.0
    for idx, r in enumerate(rows, 1):
        row_n = 3 + idx
        qty = int(r.get('quantity') or 1)
        buy = float(r.get('buy_price') or 0)
        sell = float(r.get('sell_price') or 0)
        line_sum = buy * qty
        ws.cell(row=row_n, column=1, value=idx).alignment = center
        ws.cell(row=row_n, column=2, value=r.get('title') or '').alignment = left
        ws.cell(row=row_n, column=3, value=r.get('color') or '').alignment = center
        ws.cell(row=row_n, column=4, value=qty).alignment = center
        ws.cell(row=row_n, column=5, value=int(buy)).alignment = right_a
        ws.cell(row=row_n, column=6, value=int(line_sum)).alignment = right_a
        ws.cell(row=row_n, column=7, value=r.get('sku') or '').alignment = center
        for col in range(1, 8):
            ws.cell(row=row_n, column=col).border = box
        total_qty += qty
        total_buy += line_sum
        total_sell += sell * qty

    total_row = 3 + len(rows) + 1
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    c_label = ws.cell(row=total_row, column=1, value='ИТОГО:')
    c_label.font = bold
    c_label.alignment = right_a
    c_label.border = box
    qty_cell = ws.cell(row=total_row, column=4, value=total_qty)
    qty_cell.font = bold
    qty_cell.alignment = center
    qty_cell.border = box
    sum_label = ws.cell(row=total_row, column=5, value='')
    sum_label.border = box
    sum_cell = ws.cell(row=total_row, column=6, value=int(total_buy))
    sum_cell.font = bold
    sum_cell.alignment = right_a
    sum_cell.border = box
    ws.cell(row=total_row, column=7, value='').border = box

    # Краткая сводка под таблицей
    summary_row = total_row + 2
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=7)
    ws.cell(row=summary_row, column=1, value=(
        f'Позиций: {len(rows)} · Штук всего: {total_qty} · '
        f'Закупка: {int(total_buy):,} ₽ · Розница: {int(total_sell):,} ₽'
    ).replace(',', ' ')).alignment = left

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    b64 = base64.b64encode(out.read()).decode('ascii')

    fn_parts = ['nakladnaya']
    if brand:
        # транслитерация не критична — пишем латиницей если возможно, иначе оставим
        safe = re.sub(r'[^a-zA-Z0-9_-]', '_', brand) or 'brand'
        fn_parts.append(safe)
    if date_from:
        fn_parts.append(date_from.replace('-', ''))
    fn = '_'.join(fn_parts) + '.xlsx'

    return _ok({
        'filename': fn,
        'content_base64': b64,
        'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'positions': len(rows),
        'total_qty': total_qty,
        'total_buy': int(total_buy),
        'total_sell': int(total_sell),
    })


# ============ Specs templates / autofill ============
def autofill_specs(body):
    title = (body.get('title') or '').strip()
    if not title:
        return _err(400, 'title обязателен')
    key = _norm_key(title)
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    # 1) точное совпадение
    cur.execute(f"SELECT * FROM {SCHEMA}.slshop_specs_templates WHERE match_key={_esc(key)} ORDER BY popularity DESC LIMIT 1")
    row = cur.fetchone()
    if not row:
        # 2) частичное (LIKE %key%) — попробуем найти содержащееся
        like = f"%{key}%"
        cur.execute(f"SELECT * FROM {SCHEMA}.slshop_specs_templates WHERE match_key ILIKE {_esc(like)} OR {_esc(key)} ILIKE '%' || match_key || '%' ORDER BY popularity DESC, LENGTH(match_key) DESC LIMIT 1")
        row = cur.fetchone()
    if not row:
        # 3) fallback: возьмём последний созданный пользовательский шаблон по этому ключу-первому слову
        first = key.split(' ')[0] if key else ''
        if first:
            like = f"%{first}%"
            cur.execute(f"SELECT * FROM {SCHEMA}.slshop_specs_templates WHERE match_key ILIKE {_esc(like)} ORDER BY popularity DESC LIMIT 1")
            row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        return _ok({'found': False})
    return _ok({'found': True, 'template': dict(row)})


def save_specs_template(body):
    title = (body.get('title') or body.get('match_key') or '').strip()
    if not title:
        return _err(400, 'title/match_key обязателен')
    key = _norm_key(title)
    fields = {
        'match_key': key,
        'brand': body.get('brand'),
        'model': body.get('model') or title,
        'specs_short': body.get('specs_short'),
        'specs_full': body.get('specs_full'),
        'default_color': body.get('default_color'),
        'default_storage': body.get('default_storage'),
        'category_id': body.get('category_id'),
        'popularity': int(body.get('popularity') or 1),
        'is_builtin': False,
    }
    conn = get_conn(); cur = conn.cursor()
    # увеличиваем popularity если уже есть запись с этим match_key
    cur.execute(f"SELECT id FROM {SCHEMA}.slshop_specs_templates WHERE match_key={_esc(key)} AND is_builtin=false LIMIT 1")
    ex = cur.fetchone()
    if ex:
        sets = ', '.join([f"{k}={_esc(v)}" for k, v in fields.items() if k not in ('match_key',)])
        cur.execute(f"UPDATE {SCHEMA}.slshop_specs_templates SET {sets}, popularity=popularity+1 WHERE id={ex[0]} RETURNING id")
        nid = cur.fetchone()[0]
    else:
        cols = ', '.join(fields.keys())
        vals = ', '.join([_esc(v) for v in fields.values()])
        cur.execute(f"INSERT INTO {SCHEMA}.slshop_specs_templates ({cols}) VALUES ({vals}) RETURNING id")
        nid = cur.fetchone()[0]
    conn.commit(); cur.close(); conn.close()
    return _ok({'id': nid})


# ============ Items ============
ITEM_FIELDS = [
    'sku', 'category_id', 'title', 'brand', 'model', 'specs', 'specs_short', 'storage', 'color',
    'ram_gb', 'storage_gb',
    'condition', 'imei', 'serial_number', 'battery_health', 'has_box', 'has_charger', 'description',
    'buy_price', 'sell_price', 'min_price', 'status', 'source', 'consignment_percent',
    'consignment_owner_id', 'buy_client_id', 'branch_id', 'warranty_days',
    'quantity'
]


def _gen_sku():
    return 'SL' + datetime.utcnow().strftime('%y%m%d') + '-' + str(int(datetime.utcnow().timestamp()) % 100000)


def list_items(params):
    status = (params.get('status') or '').strip()
    category_id = params.get('category_id')
    q = (params.get('q') or '').strip()
    limit = int(params.get('limit') or 200)
    where = []
    if status:
        where.append(f"i.status={_esc(status)}")
    if category_id:
        where.append(f"i.category_id={int(category_id)}")
    if q:
        like = f"%{q}%"
        where.append(f"(i.title ILIKE {_esc(like)} OR i.imei ILIKE {_esc(like)} OR i.sku ILIKE {_esc(like)})")
    wsql = (' WHERE ' + ' AND '.join(where)) if where else ''
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        f"SELECT i.*, c.name AS category_name, c.icon AS category_icon, c.path AS category_path, "
        f"cl.full_name AS buy_client_name, "
        f"b.name AS branch_name, b.address AS branch_address "
        f"FROM {SCHEMA}.slshop_items i "
        f"LEFT JOIN {SCHEMA}.slshop_categories c ON c.id=i.category_id "
        f"LEFT JOIN {SCHEMA}.slshop_clients cl ON cl.id=i.buy_client_id "
        f"LEFT JOIN {SCHEMA}.slshop_branches b ON b.id=i.branch_id "
        f"{wsql} ORDER BY i.created_at DESC LIMIT {limit}"
    )
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [dict(r) for r in rows]


def _validate_fk(cur, table, value, label):
    """Проверяет существование внешней записи. Возвращает (ok, normalized_value, err)."""
    if value in (None, '', 0, '0'):
        return True, None, None
    try:
        v = int(value)
    except (TypeError, ValueError):
        return False, None, f'{label}: некорректный id'
    cur.execute(f"SELECT 1 FROM {SCHEMA}.{table} WHERE id={v} LIMIT 1")
    if not cur.fetchone():
        return False, None, f'{label}: запись с id={v} не найдена'
    return True, v, None


def create_item(body, employee):
    title = (body.get('title') or '').strip()
    if not title:
        return _err(400, 'title обязателен')
    data = {k: body.get(k) for k in ITEM_FIELDS}
    data['title'] = title
    if not data.get('sku'):
        data['sku'] = _gen_sku()
    if not data.get('status'):
        data['status'] = 'stock'
    if data.get('buy_price') in (None, ''):
        data['buy_price'] = 0
    if data.get('sell_price') in (None, ''):
        data['sell_price'] = 0
    # Количество единиц в позиции (партия аксессуаров: чехлы, стёкла и т.п.)
    try:
        qty = int(data.get('quantity') or 1)
    except (TypeError, ValueError):
        qty = 1
    if qty < 1:
        qty = 1
    data['quantity'] = qty
    conn = get_conn(); cur = conn.cursor()
    # Валидация FK — если ссылка битая, обнуляем (не падаем)
    for fk_field, fk_table, fk_label in [
        ('category_id', 'slshop_categories', 'category_id'),
        ('branch_id', 'slshop_branches', 'branch_id'),
        ('buy_client_id', 'slshop_clients', 'buy_client_id'),
        ('consignment_owner_id', 'slshop_clients', 'consignment_owner_id'),
    ]:
        ok, norm, _err_msg = _validate_fk(cur, fk_table, data.get(fk_field), fk_label)
        data[fk_field] = norm if ok else None
    cols = ', '.join(list(data.keys()) + ['created_by', 'buy_at'])
    vals = ', '.join([_esc(v) for v in data.values()] + [_esc(employee.get('full_name') if employee else None), 'NOW()'])
    cur.execute(f"INSERT INTO {SCHEMA}.slshop_items ({cols}) VALUES ({vals}) RETURNING id, sku")
    row = cur.fetchone()
    item_id = row[0]
    sku = row[1]
    # Сумма закупки = цена за штуку × количество (для партий аксессуаров)
    unit_buy_price = float(data.get('buy_price') or 0)
    total_buy = unit_buy_price * qty
    log_note = f'Цена {unit_buy_price:g} ₽' + (f' × {qty} шт = {total_buy:g} ₽' if qty > 1 else '')
    _log_event(cur, 'buy', 'item', item_id, f'Скупка: {title}' + (f' ({qty} шт)' if qty > 1 else ''),
               log_note, total_buy, data.get('branch_id'), employee)
    # Если это скупка с ценой > 0 — создадим операцию buy
    if data.get('source', 'buyout') == 'buyout' and unit_buy_price > 0:
        cur.execute(
            f"INSERT INTO {SCHEMA}.slshop_operations (op_type, item_id, client_id, amount, employee_name, status_to) "
            f"VALUES ('buy', {item_id}, {_esc(data.get('buy_client_id'))}, {_esc(total_buy)}, "
            f"{_esc(employee.get('full_name') if employee else None)}, 'stock') RETURNING id"
        )
        op_id = cur.fetchone()[0]
        cur.execute(f"UPDATE {SCHEMA}.slshop_items SET buy_operation_id={op_id} WHERE id={item_id}")
        # Авто-расход по кассе при скупке (на полную сумму партии)
        branch_id = data.get('branch_id')
        if branch_id:
            cur.execute(
                f"SELECT id, balance FROM {SCHEMA}.slshop_cash_accounts WHERE branch_id={int(branch_id)} AND is_active=true LIMIT 1"
            )
            acc = cur.fetchone()
            if acc:
                acc_id, balance = acc[0], float(acc[1] or 0)
                amt = total_buy
                new_balance = balance - amt
                reason_text = ('Скупка №' + str(item_id)) + (f' ({qty} шт × {unit_buy_price:g}₽)' if qty > 1 else '')
                cur.execute(
                    f"INSERT INTO {SCHEMA}.slshop_cash_movements (account_id, direction, amount, balance_after, category, reason, employee_name, related_op_id, related_item_id, is_auto, branch_id) "
                    f"VALUES ({acc_id}, 'out', {amt}, {new_balance}, 'Скупка товара', "
                    f"{_esc(reason_text)}, {_esc(employee.get('full_name') if employee else None)}, {op_id}, {item_id}, TRUE, {int(branch_id)})"
                )
                cur.execute(f"UPDATE {SCHEMA}.slshop_cash_accounts SET balance={new_balance} WHERE id={acc_id}")
    conn.commit(); cur.close(); conn.close()
    # Сохраняем шаблон характеристик для будущей автоподстановки
    if data.get('specs_short') or data.get('specs'):
        try:
            save_specs_template({
                'title': title, 'brand': data.get('brand'), 'model': data.get('model'),
                'specs_short': data.get('specs_short'), 'specs_full': data.get('specs'),
                'default_color': data.get('color'), 'default_storage': data.get('storage'),
                'category_id': data.get('category_id'),
            })
        except Exception:
            pass
    # Push сотрудникам о новой скупке (только если это реальная скупка с ценой)
    try:
        if data.get('source', 'buyout') == 'buyout' and unit_buy_price > 0:
            _qty_str = f" × {qty} шт" if qty > 1 else ""
            _send_push_event(
                title=f"Скупка №{item_id}",
                body=f"{title[:60]}{_qty_str} · {total_buy:g} ₽",
                url=f"/staff?tab=smartlombard&item={item_id}",
                tag=f"slshop-{item_id}",
            )
    except Exception:
        pass
    return _ok({'id': item_id, 'sku': sku})


def update_item(body, employee=None):
    item_id = body.get('id')
    if not item_id:
        return _err(400, 'id обязателен')
    data = {k: body.get(k) for k in ITEM_FIELDS if k in body}
    if not data:
        return _err(400, 'нет полей для обновления')
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT status FROM {SCHEMA}.slshop_items WHERE id={int(item_id)}")
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close(); return _err(404, 'Товар не найден')
    cur_status = row[0]
    is_owner = employee and employee.get('role') == 'owner'
    # Проданный товар: разрешаем редактировать IMEI/серийник/описание/фото всем,
    # цену продажи — только владельцу, остальные поля — только владельцу
    ALLOWED_SOLD_FIELDS = {'imei', 'serial_number', 'description', 'specs_short', 'specs', 'condition'}
    if cur_status == 'sold' and not is_owner:
        blocked = [k for k in data if k not in ALLOWED_SOLD_FIELDS]
        if blocked:
            return _err(403, f'Редактировать проданный товар может только владелец (поля: {", ".join(blocked)})')
    if cur_status == 'sold' and 'sell_price' in data and not is_owner:
        return _err(403, 'Менять цену продажи может только владелец')
    # Валидация FK
    fk_check = [
        ('category_id', 'slshop_categories'),
        ('branch_id', 'slshop_branches'),
        ('buy_client_id', 'slshop_clients'),
        ('consignment_owner_id', 'slshop_clients'),
    ]
    for fk_field, fk_table in fk_check:
        if fk_field in data:
            ok, norm, _e = _validate_fk(cur, fk_table, data[fk_field], fk_field)
            data[fk_field] = norm if ok else None
    sets = ', '.join([f"{k}={_esc(v)}" for k, v in data.items()])
    cur.execute(f"UPDATE {SCHEMA}.slshop_items SET {sets}, updated_at=NOW() WHERE id={int(item_id)} RETURNING id")
    row = cur.fetchone()
    conn.commit(); cur.close(); conn.close()
    if not row:
        return _err(404, 'Товар не найден')
    return _ok({'id': row[0]})


def items_recognize_storage(body, employee):
    """Массовое распознавание ОЗУ/Памяти из текстового поля storage у товаров.
    Парсит '4/128', '4/128GB', '128GB', '128' и заполняет ram_gb/storage_gb там, где они NULL.
    Возвращает количество обновлённых записей и примеры.
    """
    if not employee:
        return _err(401, 'Не авторизован')
    only_phones = bool(body.get('only_phones', True))
    dry_run = bool(body.get('dry_run', False))
    conn = get_conn(); cur = conn.cursor()
    where_phone = ""
    if only_phones:
        where_phone = (
            f" AND (c.slug IN ('mob_phones','techno_spark') "
            f"OR LOWER(c.name) LIKE '%телефон%' OR LOWER(c.name) LIKE '%смартфон%' "
            f"OR LOWER(i.title) LIKE '%iphone%' OR LOWER(i.title) LIKE '%phone%')"
        )
    cur.execute(
        f"""
        SELECT i.id, i.title, i.storage
        FROM {SCHEMA}.slshop_items i
        LEFT JOIN {SCHEMA}.slshop_categories c ON c.id = i.category_id
        WHERE i.storage IS NOT NULL AND i.storage <> ''
          AND (i.ram_gb IS NULL OR i.storage_gb IS NULL)
          {where_phone}
        ORDER BY i.id DESC
        LIMIT 5000
        """
    )
    rows = cur.fetchall()
    updated = 0
    samples = []
    for iid, title, storage in rows:
        s = str(storage or '')
        m = re.search(r'(\d{1,3})\s*/\s*(\d{2,4})', s)
        ram = int(m.group(1)) if m else None
        st = int(m.group(2)) if m else None
        if not st:
            m2 = re.search(r'(\d{2,4})', s)
            if m2:
                st = int(m2.group(1))
        if not ram and not st:
            continue
        if dry_run:
            updated += 1
            if len(samples) < 10:
                samples.append({'id': iid, 'title': title, 'storage': s, 'ram_gb': ram, 'storage_gb': st})
            continue
        sets = []
        if ram:
            sets.append(f"ram_gb=COALESCE(ram_gb, {ram})")
        if st:
            sets.append(f"storage_gb=COALESCE(storage_gb, {st})")
        if not sets:
            continue
        cur.execute(
            f"UPDATE {SCHEMA}.slshop_items SET {', '.join(sets)}, updated_at=NOW() "
            f"WHERE id={int(iid)} AND (ram_gb IS NULL OR storage_gb IS NULL)"
        )
        if cur.rowcount > 0:
            updated += 1
            if len(samples) < 10:
                samples.append({'id': iid, 'title': title, 'storage': s, 'ram_gb': ram, 'storage_gb': st})
    if not dry_run:
        conn.commit()
    cur.close(); conn.close()
    return _ok({'updated': updated, 'scanned': len(rows), 'dry_run': dry_run, 'samples': samples})


def remove_item(body, employee):
    """Удаление товара. Проданный — только владелец."""
    if not employee:
        return _err(401, 'Не авторизован')
    item_id = body.get('id')
    if not item_id:
        return _err(400, 'id обязателен')
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT status FROM {SCHEMA}.slshop_items WHERE id={int(item_id)}")
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close(); return _err(404, 'Товар не найден')
    cur_status = row[0]
    is_owner = employee.get('role') == 'owner'
    if cur_status == 'sold' and not is_owner:
        cur.close(); conn.close()
        return _err(403, 'Удалять проданный товар может только владелец')
    # очищаем все возможные FK-ссылки, чтобы DELETE не упал
    cur.execute(f"UPDATE {SCHEMA}.slshop_operations SET item_id=NULL WHERE item_id={int(item_id)}")
    cur.execute(f"UPDATE {SCHEMA}.slshop_cash_movements SET related_item_id=NULL WHERE related_item_id={int(item_id)}")
    cur.execute(f"UPDATE {SCHEMA}.slshop_discount_log SET item_id=NULL WHERE item_id={int(item_id)}")
    cur.execute(f"UPDATE {SCHEMA}.slshop_revision_items SET item_id=NULL WHERE item_id={int(item_id)}")
    cur.execute(f"DELETE FROM {SCHEMA}.slshop_items WHERE id={int(item_id)}")
    conn.commit(); cur.close(); conn.close()
    return _ok({'ok': True})


# ============ Roles & permissions ============
def list_roles():
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(f"SELECT * FROM {SCHEMA}.slshop_roles ORDER BY sort_order, name")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [dict(r) for r in rows]


def save_role(body, employee):
    if not employee or employee.get('role') != 'owner':
        return _err(403, 'Изменять роли может только владелец')
    rid = body.get('id')
    code = (body.get('code') or '').strip().lower()
    name = (body.get('name') or '').strip()
    description = body.get('description') or ''
    permissions = body.get('permissions') or {}
    if not name:
        return _err(400, 'Название обязательно')
    if not isinstance(permissions, dict):
        return _err(400, 'permissions должен быть объектом')
    perms_json = json.dumps(permissions, ensure_ascii=False).replace("'", "''")
    conn = get_conn(); cur = conn.cursor()
    if rid:
        cur.execute(
            f"UPDATE {SCHEMA}.slshop_roles SET name={_esc(name)}, description={_esc(description)}, "
            f"permissions='{perms_json}'::jsonb, updated_at=NOW() WHERE id={int(rid)} RETURNING id"
        )
        row = cur.fetchone()
        nid = row[0] if row else rid
    else:
        if not code:
            code = 'role_' + str(int(datetime.utcnow().timestamp()))
        cur.execute(
            f"INSERT INTO {SCHEMA}.slshop_roles (code, name, description, permissions, is_system) "
            f"VALUES ({_esc(code)}, {_esc(name)}, {_esc(description)}, '{perms_json}'::jsonb, FALSE) RETURNING id"
        )
        nid = cur.fetchone()[0]
    conn.commit(); cur.close(); conn.close()
    return _ok({'id': nid})


def my_permissions(employee):
    """Возвращает права текущего сотрудника на основе его роли"""
    if not employee:
        return _err(401, 'Не авторизован')
    role_code = employee.get('role') or 'staff'
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(f"SELECT code, name, permissions FROM {SCHEMA}.slshop_roles WHERE code={_esc(role_code)} LIMIT 1")
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        return _ok({'role': role_code, 'name': role_code, 'permissions': {}, 'is_owner': role_code == 'owner'})
    perms = row['permissions'] or {}
    return _ok({'role': row['code'], 'name': row['name'], 'permissions': perms, 'is_owner': row['code'] == 'owner'})


# ============ Cash accounts ============
def list_cash_accounts():
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        f"SELECT a.*, b.name AS branch_name FROM {SCHEMA}.slshop_cash_accounts a "
        f"LEFT JOIN {SCHEMA}.slshop_branches b ON b.id=a.branch_id "
        f"WHERE a.is_active=true ORDER BY a.is_default DESC, a.id"
    )
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [dict(r) for r in rows]


def cash_movement_create(body, employee):
    account_id = body.get('account_id')
    direction = (body.get('direction') or '').strip()
    amount = body.get('amount')
    if not account_id or direction not in ('in', 'out') or amount in (None, ''):
        return _err(400, 'account_id, direction (in/out), amount обязательны')
    amount_f = float(amount)
    if amount_f <= 0:
        return _err(400, 'Сумма должна быть больше нуля')
    category = body.get('category') or ('Поступление' if direction == 'in' else 'Расход')
    reason = body.get('reason') or ''
    taken_by = body.get('taken_by') or ''
    # Дата операции (опционально, для проведения задним числом)
    created_at_raw = body.get('created_at') or ''
    created_at_sql = 'NOW()'
    if isinstance(created_at_raw, str) and created_at_raw.strip():
        s = created_at_raw.strip().replace('T', ' ')[:19]
        safe = ''.join(ch for ch in s if ch.isdigit() or ch in '-: ')
        if len(safe) >= 10:
            created_at_sql = f"'{safe}'::timestamptz"
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT balance, branch_id FROM {SCHEMA}.slshop_cash_accounts WHERE id={int(account_id)}")
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close(); return _err(404, 'Касса не найдена')
    balance = float(row[0] or 0)
    branch_id = row[1]
    new_balance = balance + amount_f if direction == 'in' else balance - amount_f
    cur.execute(
        f"INSERT INTO {SCHEMA}.slshop_cash_movements (account_id, direction, amount, balance_after, category, reason, taken_by, employee_name, branch_id, created_at) "
        f"VALUES ({int(account_id)}, {_esc(direction)}, {amount_f}, {new_balance}, {_esc(category)}, {_esc(reason)}, {_esc(taken_by)}, "
        f"{_esc(employee.get('full_name') if employee else None)}, {_esc(branch_id)}, {created_at_sql}) RETURNING id"
    )
    mid = cur.fetchone()[0]
    cur.execute(f"UPDATE {SCHEMA}.slshop_cash_accounts SET balance={new_balance} WHERE id={int(account_id)}")
    conn.commit(); cur.close(); conn.close()
    return _ok({'id': mid, 'balance_after': new_balance})


def cash_movements_list(params):
    account_id = params.get('account_id')
    date_from = params.get('date_from')
    date_to = params.get('date_to')
    where = []
    if account_id:
        where.append(f"m.account_id={int(account_id)}")
    if date_from:
        where.append(f"m.created_at >= {_esc(date_from)}")
    if date_to:
        where.append(f"m.created_at < {_esc(date_to)}::date + INTERVAL '1 day'")
    wsql = (' WHERE ' + ' AND '.join(where)) if where else ''
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        f"SELECT m.*, a.name AS account_name, b.name AS branch_name "
        f"FROM {SCHEMA}.slshop_cash_movements m "
        f"LEFT JOIN {SCHEMA}.slshop_cash_accounts a ON a.id=m.account_id "
        f"LEFT JOIN {SCHEMA}.slshop_branches b ON b.id=m.branch_id "
        f"{wsql} ORDER BY m.created_at DESC LIMIT 300"
    )
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [dict(r) for r in rows]


def cash_summary():
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        f"SELECT a.id, a.name, a.balance, b.name AS branch_name, "
        f"(SELECT COALESCE(SUM(amount),0) FROM {SCHEMA}.slshop_cash_movements WHERE account_id=a.id AND direction='in' AND created_at::date = CURRENT_DATE) AS today_in, "
        f"(SELECT COALESCE(SUM(amount),0) FROM {SCHEMA}.slshop_cash_movements WHERE account_id=a.id AND direction='out' AND created_at::date = CURRENT_DATE) AS today_out "
        f"FROM {SCHEMA}.slshop_cash_accounts a LEFT JOIN {SCHEMA}.slshop_branches b ON b.id=a.branch_id "
        f"WHERE a.is_active=true ORDER BY a.is_default DESC, a.id"
    )
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [dict(r) for r in rows]


# ============ Requisites ============
def list_requisites():
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        f"SELECT r.*, b.name AS branch_name FROM {SCHEMA}.slshop_requisites r "
        f"LEFT JOIN {SCHEMA}.slshop_branches b ON b.id=r.branch_id "
        f"WHERE r.is_active=true ORDER BY r.is_default DESC, r.id"
    )
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [dict(r) for r in rows]


def save_requisite(body, employee):
    if not employee or employee.get('role') not in ('owner', 'manager'):
        return _err(403, 'Изменять реквизиты может только владелец или руководитель')
    rid = body.get('id')
    fields = {
        'branch_id': body.get('branch_id'),
        'legal_name': (body.get('legal_name') or '').strip(),
        'short_name': body.get('short_name'),
        'inn': body.get('inn'),
        'ogrn': body.get('ogrn'),
        'kpp': body.get('kpp'),
        'legal_address': body.get('legal_address'),
        'actual_address': body.get('actual_address'),
        'bank_name': body.get('bank_name'),
        'bank_bic': body.get('bank_bic'),
        'bank_account': body.get('bank_account'),
        'corr_account': body.get('corr_account'),
        'phone': body.get('phone'),
        'email': body.get('email'),
        'director_name': body.get('director_name'),
        'director_position': body.get('director_position'),
        'signatory_name': body.get('signatory_name'),
        'warranty_days': int(body.get('warranty_days') or 365),
        'is_default': bool(body.get('is_default', False)),
    }
    if not fields['legal_name']:
        return _err(400, 'legal_name обязателен')
    conn = get_conn(); cur = conn.cursor()
    if rid:
        sets = ', '.join([f"{k}={_esc(v)}" for k, v in fields.items()])
        cur.execute(f"UPDATE {SCHEMA}.slshop_requisites SET {sets}, updated_at=NOW() WHERE id={int(rid)} RETURNING id")
        row = cur.fetchone()
        nid = row[0] if row else rid
    else:
        cols = ', '.join(fields.keys())
        vals = ', '.join([_esc(v) for v in fields.values()])
        cur.execute(f"INSERT INTO {SCHEMA}.slshop_requisites ({cols}) VALUES ({vals}) RETURNING id")
        nid = cur.fetchone()[0]
    conn.commit(); cur.close(); conn.close()
    return _ok({'id': nid})


# ============ Document templates ============
def list_doc_templates(params):
    op_type = (params.get('op_type') or '').strip()
    only_active = params.get('only_active') == '1'
    where = []
    if only_active:
        where.append("is_active=true")
    if op_type:
        where.append(f"({_esc(op_type)} = ANY(op_types))")
    wsql = (' WHERE ' + ' AND '.join(where)) if where else ''
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(f"SELECT * FROM {SCHEMA}.slshop_doc_templates{wsql} ORDER BY sort_order, name")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [dict(r) for r in rows]


def doc_template_toggle(body, employee):
    if not employee or employee.get('role') not in ('owner', 'manager'):
        return _err(403, 'Менять шаблоны может только владелец или руководитель')
    tid = body.get('id')
    if not tid:
        return _err(400, 'id обязателен')
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"UPDATE {SCHEMA}.slshop_doc_templates SET is_active = NOT is_active WHERE id={int(tid)} RETURNING is_active")
    row = cur.fetchone()
    conn.commit(); cur.close(); conn.close()
    return _ok({'is_active': row[0] if row else None})


def doc_context(params):
    """Возвращает все данные для печати документа: товар, клиент, реквизиты, филиал, операция, сотрудник."""
    item_id = params.get('item_id')
    op_id = params.get('op_id')
    client_id = params.get('client_id')
    if not (item_id or op_id):
        return _err(400, 'нужен item_id или op_id')
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    item = None; op = None; client = None; branch = None; req = None
    if item_id:
        cur.execute(
            f"SELECT i.*, c.name AS category_name, c.path AS category_path, b.name AS branch_name, b.address AS branch_address "
            f"FROM {SCHEMA}.slshop_items i "
            f"LEFT JOIN {SCHEMA}.slshop_categories c ON c.id=i.category_id "
            f"LEFT JOIN {SCHEMA}.slshop_branches b ON b.id=i.branch_id "
            f"WHERE i.id={int(item_id)}"
        )
        r = cur.fetchone()
        if r:
            item = dict(r)
            if item.get('branch_id') and not client_id:
                client_id = item.get('buy_client_id')
    if op_id:
        cur.execute(
            f"SELECT o.*, i.title AS item_title, i.imei AS item_imei, i.sku AS item_sku, "
            f"i.specs_short, i.sell_price, i.buy_price, i.branch_id, "
            f"b.name AS branch_name, b.address AS branch_address, "
            f"c.full_name AS client_name, c.phone AS client_phone, c.passport_series, c.passport_number, "
            f"c.passport_issued_by, c.passport_issued_date, c.address AS client_address, c.birth_date "
            f"FROM {SCHEMA}.slshop_operations o "
            f"LEFT JOIN {SCHEMA}.slshop_items i ON i.id=o.item_id "
            f"LEFT JOIN {SCHEMA}.slshop_branches b ON b.id=i.branch_id "
            f"LEFT JOIN {SCHEMA}.slshop_clients c ON c.id=o.client_id "
            f"WHERE o.id={int(op_id)}"
        )
        r = cur.fetchone()
        if r:
            op = dict(r)
            if not item_id and op.get('item_id'):
                cur.execute(f"SELECT * FROM {SCHEMA}.slshop_items WHERE id={int(op['item_id'])}")
                ir = cur.fetchone()
                if ir:
                    item = dict(ir)
            if not client_id and op.get('client_id'):
                client_id = op['client_id']
    if client_id:
        cur.execute(f"SELECT * FROM {SCHEMA}.slshop_clients WHERE id={int(client_id)}")
        r = cur.fetchone()
        if r:
            client = dict(r)
    branch_id = (item or {}).get('branch_id') or (op or {}).get('branch_id')
    if branch_id:
        cur.execute(f"SELECT * FROM {SCHEMA}.slshop_branches WHERE id={int(branch_id)}")
        r = cur.fetchone()
        if r:
            branch = dict(r)
        cur.execute(
            f"SELECT * FROM {SCHEMA}.slshop_requisites WHERE branch_id={int(branch_id)} AND is_active=true LIMIT 1"
        )
        r = cur.fetchone()
        if r:
            req = dict(r)
    if not req:
        cur.execute(f"SELECT * FROM {SCHEMA}.slshop_requisites WHERE is_default=true AND is_active=true LIMIT 1")
        r = cur.fetchone()
        if r:
            req = dict(r)
    cur.close(); conn.close()
    return _ok({'item': item, 'operation': op, 'client': client, 'branch': branch, 'requisites': req})


def sell_item(body, employee):
    item_id = body.get('item_id')
    amount = body.get('amount')
    payment = body.get('payment_method') or 'cash'
    client_id = body.get('client_id')
    contract = body.get('contract_number')
    note = body.get('note')
    # Дата операции (опционально, для проведения задним числом)
    sold_at_raw = (body.get('sold_at') or '').strip() if isinstance(body.get('sold_at'), str) else ''
    # Сколько штук продаём за один раз (для партий аксессуаров). По умолчанию 1.
    try:
        sell_qty = int(body.get('quantity') or 1)
    except (TypeError, ValueError):
        sell_qty = 1
    if sell_qty < 1:
        sell_qty = 1
    if not item_id:
        return _err(400, 'item_id обязателен')
    if amount in (None, ''):
        return _err(400, 'amount обязателен')
    # Готовим SQL-выражение даты: если передали sold_at — используем его, иначе NOW()
    sold_at_sql = 'NOW()'
    if sold_at_raw:
        # формат datetime-local: '2026-05-09T14:30' или 'YYYY-MM-DD'
        s = sold_at_raw.replace('T', ' ')[:19]
        # экранирование на всякий случай (только цифры/дефисы/двоеточия/пробелы)
        safe = ''.join(ch for ch in s if ch.isdigit() or ch in '-: ')
        if len(safe) >= 10:
            sold_at_sql = f"'{safe}'::timestamptz"
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT status, COALESCE(quantity, 1) FROM {SCHEMA}.slshop_items WHERE id={int(item_id)}")
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close(); return _err(404, 'Товар не найден')
    cur_status = row[0]
    cur_qty = int(row[1] or 1)
    if cur_status == 'sold' or cur_qty <= 0:
        cur.close(); conn.close(); return _err(400, 'Товар уже продан')
    if sell_qty > cur_qty:
        cur.close(); conn.close()
        return _err(400, f'На складе осталось {cur_qty} шт, нельзя продать {sell_qty}')
    new_qty = cur_qty - sell_qty
    final_status = 'sold' if new_qty <= 0 else cur_status
    # Валидация client_id (если битый — обнулим, не падаем)
    ok_cl, client_id_norm, _err_cl = _validate_fk(cur, 'slshop_clients', client_id, 'client_id')
    client_id = client_id_norm if ok_cl else None
    cur.execute(
        f"INSERT INTO {SCHEMA}.slshop_operations (op_type, item_id, client_id, amount, payment_method, contract_number, note, employee_name, status_from, status_to, created_at) "
        f"VALUES ('sell', {int(item_id)}, {_esc(client_id)}, {_esc(amount)}, {_esc(payment)}, {_esc(contract)}, {_esc(note)}, {_esc(employee.get('full_name') if employee else None)}, {_esc(cur_status)}, {_esc(final_status)}, {sold_at_sql}) RETURNING id"
    )
    op_id = cur.fetchone()[0]
    if new_qty <= 0:
        # Партия закрыта (или это был штучный товар) — переводим в sold
        cur.execute(
            f"UPDATE {SCHEMA}.slshop_items SET status='sold', sell_price={_esc(amount)}, sell_operation_id={op_id}, "
            f"sell_at={sold_at_sql}, warranty_until=(CURRENT_DATE + COALESCE(warranty_days, 365) * INTERVAL '1 day')::date, "
            f"quantity=0, updated_at=NOW() WHERE id={int(item_id)} RETURNING branch_id, title"
        )
    else:
        # Партия частично продана — статус оставляем, уменьшаем quantity
        cur.execute(
            f"UPDATE {SCHEMA}.slshop_items SET quantity={new_qty}, sell_price={_esc(amount)}, "
            f"updated_at=NOW() WHERE id={int(item_id)} RETURNING branch_id, title"
        )
    sold_row = cur.fetchone()
    sold_branch = sold_row[0] if sold_row else None
    sold_title = sold_row[1] if sold_row and len(sold_row) > 1 else ''
    qty_note = f' ({sell_qty} шт)' if sell_qty > 1 else ''
    _log_event(cur, 'sale', 'item', int(item_id), f'Продажа: {sold_title}{qty_note}', f'Сумма {amount} ₽', float(amount), sold_branch, employee)
    # Автоматическое движение по кассе при оплате наличными
    if (payment or 'cash') == 'cash' and sold_branch:
        cur.execute(
            f"SELECT id, balance FROM {SCHEMA}.slshop_cash_accounts WHERE branch_id={int(sold_branch)} AND is_active=true LIMIT 1"
        )
        acc = cur.fetchone()
        if acc:
            acc_id, balance = acc[0], float(acc[1] or 0)
            new_balance = balance + float(amount)
            cur.execute(
                f"INSERT INTO {SCHEMA}.slshop_cash_movements (account_id, direction, amount, balance_after, category, reason, employee_name, related_op_id, related_item_id, is_auto, branch_id) "
                f"VALUES ({acc_id}, 'in', {float(amount)}, {new_balance}, 'Продажа товара', "
                f"{_esc('Продажа №' + str(item_id))}, {_esc(employee.get('full_name') if employee else None)}, {op_id}, {int(item_id)}, TRUE, {int(sold_branch)})"
            )
            cur.execute(f"UPDATE {SCHEMA}.slshop_cash_accounts SET balance={new_balance} WHERE id={acc_id}")
    conn.commit(); cur.close(); conn.close()

    # Уведомляем сотрудников о продаже в MAX и Telegram
    try:
        emp_name = (employee.get('full_name') if employee else None) or 'Сотрудник'
        _notify_sale(sold_title or 'Товар', sold_title or 'Товар', float(amount), emp_name, sell_qty)
    except Exception as ne:
        print(f'[slshop][sell_item] notify error: {ne}')

    return _ok({'op_id': op_id})


def return_item(body, employee):
    item_id = body.get('item_id')
    note = body.get('note') or 'возврат'
    if not item_id:
        return _err(400, 'item_id обязателен')
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT status FROM {SCHEMA}.slshop_items WHERE id={int(item_id)}")
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close(); return _err(404, 'Товар не найден')
    cur_status = row[0]
    cur.execute(
        f"INSERT INTO {SCHEMA}.slshop_operations (op_type, item_id, amount, note, employee_name, status_from, status_to) "
        f"VALUES ('return', {int(item_id)}, 0, {_esc(note)}, {_esc(employee.get('full_name') if employee else None)}, {_esc(cur_status)}, 'returned') RETURNING id"
    )
    op_id = cur.fetchone()[0]
    cur.execute(f"UPDATE {SCHEMA}.slshop_items SET status='returned', updated_at=NOW() WHERE id={int(item_id)}")
    conn.commit(); cur.close(); conn.close()
    return _ok({'op_id': op_id})


def change_status(body):
    item_id = body.get('item_id')
    new_status = body.get('status')
    if not item_id or not new_status:
        return _err(400, 'item_id и status обязательны')
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT status FROM {SCHEMA}.slshop_items WHERE id={int(item_id)}")
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close(); return _err(404, 'Товар не найден')
    cur_status = row[0]
    cur.execute(f"UPDATE {SCHEMA}.slshop_items SET status={_esc(new_status)}, updated_at=NOW() WHERE id={int(item_id)}")
    cur.execute(
        f"INSERT INTO {SCHEMA}.slshop_operations (op_type, item_id, status_from, status_to) "
        f"VALUES ('move', {int(item_id)}, {_esc(cur_status)}, {_esc(new_status)})"
    )
    conn.commit(); cur.close(); conn.close()
    return _ok({'ok': True})


# ============ Operations ============
def list_operations(params):
    op_type = (params.get('op_type') or '').strip()
    date_from = (params.get('date_from') or '').strip()
    date_to = (params.get('date_to') or '').strip()
    where = []
    if op_type:
        where.append(f"o.op_type={_esc(op_type)}")
    if date_from:
        where.append(f"o.created_at>={_esc(date_from)}")
    if date_to:
        where.append(f"o.created_at<={_esc(date_to)} + INTERVAL '1 day'")
    wsql = (' WHERE ' + ' AND '.join(where)) if where else ''
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        f"SELECT o.*, i.title AS item_title, i.imei AS item_imei, c.full_name AS client_name "
        f"FROM {SCHEMA}.slshop_operations o "
        f"LEFT JOIN {SCHEMA}.slshop_items i ON i.id=o.item_id "
        f"LEFT JOIN {SCHEMA}.slshop_clients c ON c.id=o.client_id "
        f"{wsql} ORDER BY o.created_at DESC LIMIT 300"
    )
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [dict(r) for r in rows]


# ============ Stats ============
def stats(params):
    period = (params.get('period') or '30d').strip()
    now = datetime.utcnow()
    # Произвольный период (конкретная дата или диапазон): date_from + date_to из query
    if period == 'custom':
        df_raw = (params.get('date_from') or '').strip()
        dt_raw = (params.get('date_to') or '').strip()
        # валидируем формат YYYY-MM-DD
        def _safe_date(s):
            try:
                datetime.strptime(s[:10], '%Y-%m-%d')
                return s[:10]
            except Exception:
                return None
        df_safe = _safe_date(df_raw)
        dt_safe = _safe_date(dt_raw)
        if df_safe and dt_safe:
            date_from = df_safe
            date_to = dt_safe
        else:
            # fallback: если дата невалидная, берём 30 дней
            date_from = (now - timedelta(days=30)).strftime('%Y-%m-%d')
            date_to = now.strftime('%Y-%m-%d')
    elif period == 'today':
        date_from = now.strftime('%Y-%m-%d')
    elif period == 'yesterday':
        date_from = (now - timedelta(days=1)).strftime('%Y-%m-%d')
        date_to = (now - timedelta(days=1)).strftime('%Y-%m-%d')
    elif period == '7d':
        date_from = (now - timedelta(days=7)).strftime('%Y-%m-%d')
    elif period == 'year':
        date_from = (now - timedelta(days=365)).strftime('%Y-%m-%d')
    elif period == 'all':
        date_from = '2000-01-01'
    else:
        date_from = (now - timedelta(days=30)).strftime('%Y-%m-%d')
    if 'date_to' not in dir():
        date_to = now.strftime('%Y-%m-%d')
    cond = f" WHERE created_at>={_esc(date_from)} AND created_at<{_esc(date_to)}::date + INTERVAL '1 day'"
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    # купили/потратили
    cur.execute(f"SELECT COUNT(*) c, COALESCE(SUM(amount),0) s FROM {SCHEMA}.slshop_operations{cond} AND op_type='buy'")
    bought = cur.fetchone()
    # продали/выручка
    cur.execute(f"SELECT COUNT(*) c, COALESCE(SUM(amount),0) s FROM {SCHEMA}.slshop_operations{cond} AND op_type='sell'")
    sold = cur.fetchone()
    # возвраты
    cur.execute(f"SELECT COUNT(*) c FROM {SCHEMA}.slshop_operations{cond} AND op_type='return'")
    returns = cur.fetchone()
    # склад/витрина/реализация
    cur.execute(
        f"SELECT status, COUNT(*) c, COALESCE(SUM(sell_price),0) s "
        f"FROM {SCHEMA}.slshop_items GROUP BY status"
    )
    by_status = {r['status']: {'count': int(r['c']), 'sum': float(r['s'])} for r in cur.fetchall()}
    # продажи по категориям
    cur.execute(
        f"SELECT c.name, COUNT(*) cnt, COALESCE(SUM(o.amount),0) s "
        f"FROM {SCHEMA}.slshop_operations o "
        f"JOIN {SCHEMA}.slshop_items i ON i.id=o.item_id "
        f"LEFT JOIN {SCHEMA}.slshop_categories c ON c.id=i.category_id "
        f"WHERE o.op_type='sell' AND o.created_at>={_esc(date_from)} "
        f"GROUP BY c.name ORDER BY s DESC"
    )
    by_category = [dict(r) for r in cur.fetchall()]
    # топ моделей
    cur.execute(
        f"SELECT i.title, COUNT(*) cnt, COALESCE(SUM(o.amount),0) s "
        f"FROM {SCHEMA}.slshop_operations o "
        f"JOIN {SCHEMA}.slshop_items i ON i.id=o.item_id "
        f"WHERE o.op_type='sell' AND o.created_at>={_esc(date_from)} "
        f"GROUP BY i.title ORDER BY cnt DESC LIMIT 10"
    )
    top_models = [dict(r) for r in cur.fetchall()]
    # графики по дням
    cur.execute(
        f"SELECT DATE(created_at) d, op_type, COUNT(*) cnt, COALESCE(SUM(amount),0) s "
        f"FROM {SCHEMA}.slshop_operations "
        f"WHERE created_at>={_esc(date_from)} AND op_type IN ('buy','sell') "
        f"GROUP BY 1, 2 ORDER BY 1"
    )
    daily = [dict(r) for r in cur.fetchall()]
    # Себестоимость проданного за период (COGS) — для корректной прибыли
    cur.execute(
        f"SELECT COALESCE(SUM(COALESCE(i.buy_price, 0)), 0) AS cogs "
        f"FROM {SCHEMA}.slshop_operations o "
        f"LEFT JOIN {SCHEMA}.slshop_items i ON i.id = o.item_id "
        f"WHERE o.op_type='sell' AND o.created_at>={_esc(date_from)} "
        f"AND o.created_at<{_esc(date_to)}::date + INTERVAL '1 day'"
    )
    cogs_row = cur.fetchone() or {}
    cogs = float(cogs_row.get('cogs') or 0)
    # СмартЛомбард: детализация по закрытым договорам за период
    # Прибыль = paid_total - amount (то, что заплатил клиент - тело займа)
    contract_closed_count = 0
    contract_issued = 0.0       # сколько денег выдали клиентам по этим договорам (тело)
    contract_returned = 0.0     # сколько вернули (paid_total)
    contract_profit = 0.0       # прибыль = returned - issued
    contract_active_count = 0
    contract_active_issued = 0.0
    try:
        # Закрытые договоры за период
        cur.execute(
            f"SELECT COUNT(*) AS cnt, "
            f"COALESCE(SUM(amount),0) AS issued, "
            f"COALESCE(SUM(paid_total),0) AS returned "
            f"FROM {SCHEMA}.contracts_14d "
            f"WHERE status='closed' "
            f"AND closed_at>={_esc(date_from)} "
            f"AND closed_at<{_esc(date_to)}::date + INTERVAL '1 day'"
        )
        r = cur.fetchone() or {}
        contract_closed_count = int(r.get('cnt') or 0)
        contract_issued = float(r.get('issued') or 0)
        contract_returned = float(r.get('returned') or 0)
        contract_profit = contract_returned - contract_issued
        # Активные договоры (всего сейчас, не за период)
        cur.execute(
            f"SELECT COUNT(*) AS cnt, COALESCE(SUM(amount),0) AS issued "
            f"FROM {SCHEMA}.contracts_14d WHERE status='active'"
        )
        r2 = cur.fetchone() or {}
        contract_active_count = int(r2.get('cnt') or 0)
        contract_active_issued = float(r2.get('issued') or 0)
    except Exception:
        pass
    # Список проданных позиций за период (для модалки с деталями)
    sold_items = []
    try:
        cur.execute(
            f"SELECT o.id AS op_id, o.item_id, o.amount AS sell_amount, o.created_at AS sold_at, "
            f"i.title, i.brand, i.model, i.sku, COALESCE(i.buy_price, 0) AS buy_price "
            f"FROM {SCHEMA}.slshop_operations o "
            f"LEFT JOIN {SCHEMA}.slshop_items i ON i.id = o.item_id "
            f"WHERE o.op_type='sell' AND o.created_at>={_esc(date_from)} "
            f"AND o.created_at<{_esc(date_to)}::date + INTERVAL '1 day' "
            f"ORDER BY o.created_at DESC LIMIT 200"
        )
        for r in cur.fetchall():
            sell = float(r.get('sell_amount') or 0)
            buy = float(r.get('buy_price') or 0)
            title = r.get('title') or ''
            brand = r.get('brand') or ''
            model_v = r.get('model') or ''
            name = title or ' '.join(x for x in [brand, model_v] if x).strip() or 'Товар'
            sold_items.append({
                'id': r.get('op_id'),
                'item_id': r.get('item_id'),
                'name': name,
                'sku': r.get('sku') or '',
                'sell_price': sell,
                'buy_price': buy,
                'profit': sell - buy,
                'sold_at': r['sold_at'].isoformat() if r.get('sold_at') else None,
            })
    except Exception:
        pass
    # Список купленных позиций (скупка) за период
    bought_items = []
    try:
        cur.execute(
            f"SELECT o.id AS op_id, o.item_id, o.amount AS buy_amount, o.created_at AS bought_at, "
            f"i.title, i.brand, i.model, i.sku "
            f"FROM {SCHEMA}.slshop_operations o "
            f"LEFT JOIN {SCHEMA}.slshop_items i ON i.id = o.item_id "
            f"WHERE o.op_type='buy' AND o.created_at>={_esc(date_from)} "
            f"AND o.created_at<{_esc(date_to)}::date + INTERVAL '1 day' "
            f"ORDER BY o.created_at DESC LIMIT 200"
        )
        for r in cur.fetchall():
            amt = float(r.get('buy_amount') or 0)
            title = r.get('title') or ''
            brand = r.get('brand') or ''
            model_v = r.get('model') or ''
            name = title or ' '.join(x for x in [brand, model_v] if x).strip() or 'Товар'
            bought_items.append({
                'id': r.get('op_id'),
                'item_id': r.get('item_id'),
                'name': name,
                'sku': r.get('sku') or '',
                'buy_price': amt,
                'bought_at': r['bought_at'].isoformat() if r.get('bought_at') else None,
            })
    except Exception:
        pass
    cur.close(); conn.close()
    revenue = float(sold['s'])
    spent = float(bought['s'])
    # Прибыль б/у — ЧИСТО от продаж: выручка − себестоимость проданного.
    # Закупки в склад (purchases) НЕ вычитаем — это инвестиция, не расход.
    profit_buy_used = revenue - cogs
    # Полная прибыль кассы за период = б/у + СмартЛомбард
    profit = profit_buy_used + contract_profit
    return _ok({
        'period': period,
        'date_from': date_from,
        'date_to': date_to,
        'bought_count': int(bought['c']),
        'spent': spent,
        'cogs': cogs,
        'sold_count': int(sold['c']),
        'revenue': revenue,
        'profit': profit,                       # общая прибыль (б/у + ломбард)
        'profit_used': profit_buy_used,         # прибыль только от б/у
        # СмартЛомбард — детализация
        'contract_closed_count': contract_closed_count,
        'contract_issued': contract_issued,
        'contract_returned': contract_returned,
        'contract_profit': contract_profit,
        'contract_active_count': contract_active_count,
        'contract_active_issued': contract_active_issued,
        'returns_count': int(returns['c']),
        'by_status': by_status,
        'by_category': by_category,
        'top_models': top_models,
        'daily': daily,
        'sold_items': sold_items,
        'bought_items': bought_items,
    })


# ============ Label templates ============
def list_label_templates():
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(f"SELECT * FROM {SCHEMA}.slshop_label_templates ORDER BY is_default DESC, id")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [dict(r) for r in rows]


def save_label_template(body):
    tid = body.get('id')
    fields = {
        'name': body.get('name') or 'Шаблон',
        'width_mm': body.get('width_mm') or 58,
        'height_mm': body.get('height_mm') or 40,
        'layout': body.get('layout') or 'classic',
        'show_brand': body.get('show_brand', True),
        'show_specs': body.get('show_specs', True),
        'show_imei': body.get('show_imei', False),
        'show_qr': body.get('show_qr', False),
        'show_barcode': body.get('show_barcode', False),
        'font_family': body.get('font_family') or 'Arial',
        'is_thermal': body.get('is_thermal', True),
    }
    conn = get_conn(); cur = conn.cursor()
    if tid:
        sets = ', '.join([f"{k}={_esc(v)}" for k, v in fields.items()])
        cur.execute(f"UPDATE {SCHEMA}.slshop_label_templates SET {sets} WHERE id={int(tid)} RETURNING id")
        row = cur.fetchone()
        nid = row[0] if row else tid
    else:
        cols = ', '.join(fields.keys())
        vals = ', '.join([_esc(v) for v in fields.values()])
        cur.execute(f"INSERT INTO {SCHEMA}.slshop_label_templates ({cols}) VALUES ({vals}) RETURNING id")
        nid = cur.fetchone()[0]
    conn.commit(); cur.close(); conn.close()
    return _ok({'id': nid})


# ============ Import / Export ============
def export_items(params):
    fmt = (params.get('format') or 'csv').lower()
    status = (params.get('status') or '').strip()
    where = ''
    if status:
        where = f" WHERE i.status={_esc(status)}"
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        f"SELECT i.sku, c.name AS category, i.title, i.brand, i.model, i.specs_short, i.specs, "
        f"i.storage, i.color, i.condition, i.imei, i.serial_number, i.battery_health, "
        f"i.buy_price, i.sell_price, i.status, i.created_at "
        f"FROM {SCHEMA}.slshop_items i LEFT JOIN {SCHEMA}.slshop_categories c ON c.id=i.category_id "
        f"{where} ORDER BY i.created_at DESC"
    )
    rows = cur.fetchall()
    cur.close(); conn.close()
    if fmt == 'json':
        return _ok({'items': [dict(r) for r in rows]})
    if fmt == 'text':
        lines = []
        for r in rows:
            d = dict(r)
            lines.append(f"{d.get('title','')} | {d.get('specs_short') or ''} | {d.get('sell_price') or 0}₽")
        return {'statusCode': 200, 'headers': {**HEADERS, 'Content-Type': 'text/plain; charset=utf-8'}, 'body': '\n'.join(lines)}
    # csv default
    buf = io.StringIO()
    if rows:
        w = csv.DictWriter(buf, fieldnames=list(dict(rows[0]).keys()), delimiter=';')
        w.writeheader()
        for r in rows:
            d = {k: ('' if v is None else v) for k, v in dict(r).items()}
            w.writerow(d)
    else:
        buf.write('sku;category;title;brand;model;specs_short;specs;storage;color;condition;imei;serial_number;battery_health;buy_price;sell_price;status;created_at\n')
    return {'statusCode': 200, 'headers': {**HEADERS, 'Content-Type': 'text/csv; charset=utf-8'}, 'body': buf.getvalue()}


def import_items(body, employee):
    rows = body.get('rows') or []
    if not isinstance(rows, list) or not rows:
        return _err(400, 'rows должен быть непустым массивом')
    created = 0
    errors = []
    conn = get_conn(); cur = conn.cursor()
    # карта категорий
    cur.execute(f"SELECT id, name, slug FROM {SCHEMA}.slshop_categories")
    cats = cur.fetchall()
    cat_map = {}
    for cid, cname, cslug in cats:
        cat_map[(cname or '').lower()] = cid
        cat_map[(cslug or '').lower()] = cid
    for idx, r in enumerate(rows):
        try:
            title = (r.get('title') or '').strip()
            if not title:
                errors.append({'row': idx, 'err': 'нет title'}); continue
            cat_id = r.get('category_id')
            if not cat_id and r.get('category'):
                cat_id = cat_map.get(str(r.get('category')).lower())
            sku = r.get('sku') or _gen_sku()
            data = {
                'sku': sku, 'category_id': cat_id, 'title': title,
                'brand': r.get('brand'), 'model': r.get('model'),
                'specs_short': r.get('specs_short'), 'specs': r.get('specs'),
                'storage': r.get('storage'), 'color': r.get('color'),
                'condition': r.get('condition'), 'imei': r.get('imei'),
                'serial_number': r.get('serial_number'),
                'buy_price': r.get('buy_price') or 0,
                'sell_price': r.get('sell_price') or 0,
                'status': r.get('status') or 'stock',
            }
            cols = ', '.join(list(data.keys()) + ['created_by', 'buy_at'])
            vals = ', '.join([_esc(v) for v in data.values()] + [_esc(employee.get('full_name') if employee else 'import'), 'NOW()'])
            cur.execute(f"INSERT INTO {SCHEMA}.slshop_items ({cols}) VALUES ({vals})")
            created += 1
        except Exception as e:
            errors.append({'row': idx, 'err': str(e)})
            try:
                conn.rollback()
            except Exception:
                pass
    conn.commit(); cur.close(); conn.close()
    return _ok({'created': created, 'errors': errors})


# ============ Router ============
def handler(event: dict, context) -> dict:
    """Единая точка входа SmartLombard (комиссионка): товары, операции, клиенты, статистика, ценники, импорт/экспорт"""
    if event.get('httpMethod') == 'OPTIONS':
        return {'statusCode': 200, 'headers': HEADERS, 'body': ''}

    method = event.get('httpMethod', 'GET')
    params = event.get('queryStringParameters') or {}
    raw = event.get('body') or '{}'
    body = json.loads(raw) if isinstance(raw, str) and raw else (raw or {})
    headers_in = {k.lower(): v for k, v in (event.get('headers') or {}).items()}
    token = headers_in.get('x-employee-token', '').strip()

    employee = get_employee_by_token(token) if token else None
    if not employee:
        return _err(401, 'Требуется авторизация сотрудника')

    action = params.get('action') or body.get('action') or ''

    try:
        # справочники
        if action == 'categories':
            return _ok(list_categories())
        if action == 'category_create' and method == 'POST':
            return create_category(body)
        if action == 'label_templates':
            return _ok(list_label_templates())
        if action == 'label_template_save' and method == 'POST':
            return save_label_template(body)
        # клиенты
        if action == 'clients':
            return _ok(list_clients(params))
        if action == 'client_save' and method == 'POST':
            return upsert_client(body)
        # шаблоны характеристик
        if action == 'autofill_specs' and method == 'POST':
            return autofill_specs(body)
        if action == 'save_specs' and method == 'POST':
            return save_specs_template(body)
        # товары
        if action == 'items':
            return _ok(list_items(params))
        if action == 'item_create' and method == 'POST':
            return create_item(body, employee)
        if action == 'item_update' and method == 'POST':
            return update_item(body, employee)
        if action == 'item_remove' and method == 'POST':
            return remove_item(body, employee)
        if action == 'items_recognize_storage' and method == 'POST':
            return items_recognize_storage(body, employee)
        # ── Накладная: распознать фото / массово создать / выгрузить XLSX ──
        if action == 'items_invoice_recognize' and method == 'POST':
            return items_invoice_recognize(body, employee)
        if action == 'items_bulk_create' and method == 'POST':
            return items_bulk_create(body, employee)
        if action == 'items_invoice_export':
            return items_invoice_export(params)
        if action == 'item_sell' and method == 'POST':
            return sell_item(body, employee)
        if action == 'item_return' and method == 'POST':
            return return_item(body, employee)
        if action == 'item_status' and method == 'POST':
            return change_status(body)
        # операции
        if action == 'operations':
            return _ok(list_operations(params))
        # статистика
        if action == 'stats':
            return stats(params)
        # импорт / экспорт
        if action == 'export':
            return export_items(params)
        if action == 'import' and method == 'POST':
            return import_items(body, employee)
        # уценка
        if action == 'discount_rules':
            return _ok(list_discount_rules())
        if action == 'discount_rule_save' and method == 'POST':
            return save_discount_rule(body)
        if action == 'discount_rule_toggle' and method == 'POST':
            return discount_rule_toggle(body)
        if action == 'discount_rule_apply' and method == 'POST':
            return discount_rule_apply(body)
        # ревизия
        if action == 'revisions':
            return _ok(list_revisions())
        if action == 'revision_create' and method == 'POST':
            return revision_create(body, employee)
        if action == 'revision_get':
            return revision_get(params)
        if action == 'revision_scan' and method == 'POST':
            return revision_scan(body, employee)
        if action == 'revision_finish' and method == 'POST':
            return revision_finish(body)
        # филиалы
        if action == 'branches':
            return _ok(list_branches())
        # удаление операций (только владелец)
        if action == 'operation_delete' and method == 'POST':
            return operation_delete(body, employee)
        # проданные товары для сводки и чеков
        if action == 'sold':
            return _ok(list_sold(params))
        # купленные товары (скупка) для сводки
        if action == 'bought':
            return _ok(list_bought(params))
        # роли и права
        if action == 'roles':
            return _ok(list_roles())
        if action == 'role_save' and method == 'POST':
            return save_role(body, employee)
        if action == 'my_permissions':
            return my_permissions(employee)
        # касса наличных
        if action == 'cash_accounts':
            return _ok(list_cash_accounts())
        if action == 'cash_summary':
            return _ok(cash_summary())
        if action == 'cash_movements':
            return _ok(cash_movements_list(params))
        if action == 'cash_movement_create' and method == 'POST':
            return cash_movement_create(body, employee)
        # реквизиты и документы
        if action == 'requisites':
            return _ok(list_requisites())
        if action == 'requisite_save' and method == 'POST':
            return save_requisite(body, employee)
        if action == 'doc_templates':
            return _ok(list_doc_templates(params))
        if action == 'doc_template_toggle' and method == 'POST':
            return doc_template_toggle(body, employee)
        if action == 'doc_context':
            return doc_context(params)
        # журнал, аналитика, бухгалтерия, избранное
        if action == 'events':
            return _ok(list_events(params))
        if action == 'analytics_full':
            return analytics_full(params)
        if action == 'inventory_report':
            return inventory_report(params)
        if action == 'accounting':
            return accounting_summary(params)
        if action == 'favorites':
            return _ok(list_favorites(employee))
        if action == 'favorite_add' and method == 'POST':
            return favorite_add(body, employee)
        if action == 'favorite_remove' and method == 'POST':
            return favorite_remove(body, employee)
        # массовый перенос товаров между филиалами
        if action == 'items_move_branch' and method == 'POST':
            return items_move_branch(body, employee)
        # фото паспорта клиента
        if action == 'client_passport_upload' and method == 'POST':
            return client_passport_upload(body, employee)
        # сотрудники / роли
        if action == 'employees_list':
            return employees_list_full(employee)
        if action == 'employee_set_role' and method == 'POST':
            return employee_set_role(body, employee)
        return _err(400, f'Неизвестный action: {action}')
    except psycopg2.errors.ForeignKeyViolation as e:
        msg = str(e).split('\n')[0] if e else 'foreign key constraint violation'
        # Извлечь имя constraint и таблицы для понятности
        import re as _re
        m = _re.search(r'on table "([^"]+)".*constraint "([^"]+)"', str(e))
        if m:
            return _err(400, f'Невозможно удалить — есть связанные записи в таблице {m.group(1)} (constraint {m.group(2)}). action={action}')
        return _err(400, f'Связанная запись не найдена. {msg} (action={action})')
    except psycopg2.errors.InsufficientPrivilege as e:
        msg = str(e).split('\n')[0] if e else 'insufficient privilege'
        return _err(403, f'Нет прав в БД для этой операции. {msg} (action={action})')
    except psycopg2.Error as e:
        # Любая другая ошибка psycopg
        msg = str(e).split('\n')[0] if e else str(e)
        return _err(500, f'БД ошибка ({action}): {type(e).__name__}: {msg}')
    except Exception as e:
        import traceback
        tb = traceback.format_exc().splitlines()[-1] if traceback else ''
        return _err(500, f'Ошибка ({action}): {type(e).__name__}: {e} | {tb}')