import json
import os
import csv
import io
import base64
from datetime import datetime

import psycopg2
from openpyxl import Workbook, load_workbook

HEADERS = {
    'Access-Control-Allow-Origin': '*',
    'Access-Control-Allow-Methods': 'GET, POST, OPTIONS',
    'Access-Control-Allow-Headers': 'Content-Type, X-Employee-Token',
    'Content-Type': 'application/json',
}
SCHEMA = 't_p31606708_tech_buying_service'


def get_conn():
    return psycopg2.connect(os.environ['DATABASE_URL'])


def auth_employee(token: str):
    if not token:
        return None
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute(
            f"SELECT id, full_name, role FROM {SCHEMA}.employees "
            f"WHERE auth_token=%s AND token_expires_at>NOW() AND is_active=true",
            (token,),
        )
        row = cur.fetchone()
        if not row:
            return None
        return {'id': row[0], 'full_name': row[1], 'role': row[2]}
    finally:
        cur.close(); conn.close()


def _ok(data, status=200):
    return {'statusCode': status, 'headers': HEADERS, 'body': json.dumps(data, ensure_ascii=False, default=str)}


def _err(status, msg):
    return {'statusCode': status, 'headers': HEADERS, 'body': json.dumps({'error': msg}, ensure_ascii=False)}


COLUMNS = [
    ('id', 'ID'),
    ('title', 'Наименование'),
    ('brand', 'Бренд'),
    ('model', 'Модель'),
    ('specs', 'Характеристики'),
    ('condition', 'Состояние'),
    ('color', 'Цвет'),
    ('storage', 'Память'),
    ('imei', 'IMEI'),
    ('serial_number', 'Серийный'),
    ('purchase_price', 'Цена закупки'),
    ('sell_price', 'Цена продажи'),
    ('status', 'Статус'),
    ('location', 'Локация'),
    ('category_name', 'Категория'),
    ('purchase_date', 'Дата закупки'),
    ('sell_date', 'Дата продажи'),
]


def fetch_items(params):
    status = params.get('status') or ''
    conn = get_conn(); cur = conn.cursor()
    try:
        where = ''; values = []
        if status:
            where = "WHERE i.status=%s"; values.append(status)
        cur.execute(
            f"SELECT i.id, i.title, i.brand, i.model, i.specs, i.condition, i.color, i.storage, "
            f"i.imei, i.serial_number, i.purchase_price, i.sell_price, i.status, i.location, "
            f"c.name AS category_name, i.purchase_date, i.sell_date "
            f"FROM {SCHEMA}.sl_items i "
            f"LEFT JOIN {SCHEMA}.sl_categories c ON c.id=i.category_id "
            f"{where} ORDER BY i.id DESC",
            values,
        )
        rows = cur.fetchall()
        keys = [c[0] for c in COLUMNS]
        return [dict(zip(keys, [str(v) if v is not None else '' for v in r])) for r in rows]
    finally:
        cur.close(); conn.close()


def export_csv(params):
    items = fetch_items(params)
    out = io.StringIO()
    writer = csv.writer(out, delimiter=';')
    writer.writerow([c[1] for c in COLUMNS])
    for it in items:
        writer.writerow([it.get(k, '') for k, _ in COLUMNS])
    csv_text = '\ufeff' + out.getvalue()  # BOM для Excel
    return _ok({'filename': f'smartlombard_{datetime.now().strftime("%Y%m%d_%H%M")}.csv',
                'content': csv_text, 'mime': 'text/csv'})


def export_text(params):
    items = fetch_items(params)
    lines = []
    for it in items:
        line = f"{it.get('title','')}"
        if it.get('specs'): line += f" — {it['specs']}"
        if it.get('storage'): line += f", {it['storage']}"
        if it.get('color'): line += f", {it['color']}"
        if it.get('condition'): line += f" ({it['condition']})"
        if it.get('sell_price') and it['sell_price'] != '0':
            line += f" — {it['sell_price']} ₽"
        lines.append(line)
    return _ok({'filename': f'smartlombard_{datetime.now().strftime("%Y%m%d_%H%M")}.txt',
                'content': '\n'.join(lines), 'mime': 'text/plain'})


def export_xlsx(params):
    items = fetch_items(params)
    wb = Workbook(); ws = wb.active; ws.title = 'Товары'
    ws.append([c[1] for c in COLUMNS])
    for it in items:
        ws.append([it.get(k, '') for k, _ in COLUMNS])
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    b64 = base64.b64encode(buf.read()).decode('ascii')
    return _ok({'filename': f'smartlombard_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
                'content_base64': b64,
                'mime': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'})


def import_items(body, employee_name):
    """Импорт списка товаров. body['items'] = массив объектов с полями
    title, brand, model, specs, condition, storage, color, imei, purchase_price,
    sell_price, category_slug."""
    items = body.get('items') or []
    if not items:
        return _err(400, 'Пустой список товаров')
    conn = get_conn(); cur = conn.cursor()
    inserted = 0; errors = []
    try:
        # карта slug → id
        cur.execute(f"SELECT id, slug FROM {SCHEMA}.sl_categories")
        cat_map = {r[1]: r[0] for r in cur.fetchall()}
        for idx, it in enumerate(items):
            try:
                title = (it.get('title') or '').strip()
                if not title:
                    errors.append({'row': idx + 1, 'error': 'нет наименования'})
                    continue
                cat_id = cat_map.get(it.get('category_slug') or 'other')
                cur.execute(
                    f"INSERT INTO {SCHEMA}.sl_items (category_id, title, brand, model, specs, condition, "
                    f"color, storage, imei, purchase_price, sell_price, status, location, source, purchase_employee) "
                    f"VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'in_stock','showcase','import',%s)",
                    (
                        cat_id, title, it.get('brand'), it.get('model'),
                        it.get('specs'), it.get('condition') or 'хорошее',
                        it.get('color'), it.get('storage'), it.get('imei'),
                        float(it.get('purchase_price') or 0),
                        float(it.get('sell_price') or 0),
                        employee_name,
                    ),
                )
                inserted += 1
            except Exception as e:
                errors.append({'row': idx + 1, 'error': str(e)})
        conn.commit()
        return _ok({'inserted': inserted, 'errors': errors, 'total': len(items)})
    finally:
        cur.close(); conn.close()


def parse_xlsx(body):
    """Парсит base64 содержимое xlsx и возвращает массив объектов для предпросмотра."""
    b64 = body.get('content_base64') or ''
    if not b64:
        return _err(400, 'content_base64 обязателен')
    raw = base64.b64decode(b64)
    wb = load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return _ok({'items': []})
    headers = [str(h or '').strip().lower() for h in rows[0]]
    # сопоставление русских и английских заголовков
    aliases = {
        'наименование': 'title', 'товар': 'title', 'name': 'title', 'title': 'title',
        'бренд': 'brand', 'brand': 'brand',
        'модель': 'model', 'model': 'model',
        'характеристики': 'specs', 'specs': 'specs',
        'состояние': 'condition', 'condition': 'condition',
        'цвет': 'color', 'color': 'color',
        'память': 'storage', 'storage': 'storage', 'memory': 'storage',
        'imei': 'imei',
        'цена закупки': 'purchase_price', 'закупка': 'purchase_price', 'purchase_price': 'purchase_price',
        'цена продажи': 'sell_price', 'цена': 'sell_price', 'sell_price': 'sell_price', 'price': 'sell_price',
        'категория': 'category_slug', 'category': 'category_slug',
    }
    field_idx = {}
    for i, h in enumerate(headers):
        key = aliases.get(h)
        if key:
            field_idx[key] = i
    if 'title' not in field_idx:
        return _err(400, 'Не найдена колонка "Наименование"/"Title"')
    items = []
    for r in rows[1:]:
        if not r or all(c is None or str(c).strip() == '' for c in r):
            continue
        obj = {}
        for k, idx in field_idx.items():
            v = r[idx] if idx < len(r) else None
            obj[k] = str(v).strip() if v is not None else ''
        if obj.get('title'):
            items.append(obj)
    return _ok({'items': items, 'count': len(items)})


def parse_csv(body):
    text = body.get('content') or ''
    if not text:
        return _err(400, 'content обязателен')
    if text.startswith('\ufeff'):
        text = text[1:]
    sample = text[:1000]
    delimiter = ';' if sample.count(';') >= sample.count(',') else ','
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    aliases = {
        'наименование': 'title', 'товар': 'title', 'name': 'title', 'title': 'title',
        'бренд': 'brand', 'brand': 'brand',
        'модель': 'model', 'model': 'model',
        'характеристики': 'specs', 'specs': 'specs',
        'состояние': 'condition', 'condition': 'condition',
        'цвет': 'color', 'color': 'color',
        'память': 'storage', 'storage': 'storage',
        'imei': 'imei',
        'цена закупки': 'purchase_price', 'purchase_price': 'purchase_price',
        'цена продажи': 'sell_price', 'цена': 'sell_price', 'price': 'sell_price',
        'категория': 'category_slug',
    }
    items = []
    for r in reader:
        obj = {}
        for k, v in r.items():
            key = aliases.get((k or '').strip().lower())
            if key:
                obj[key] = (v or '').strip()
        if obj.get('title'):
            items.append(obj)
    return _ok({'items': items, 'count': len(items)})


def handler(event: dict, context) -> dict:
    """Импорт и экспорт товаров СмартЛомбард в CSV/XLSX/текст."""
    if event.get('httpMethod') == 'OPTIONS':
        return {'statusCode': 200, 'headers': HEADERS, 'body': ''}

    method = event.get('httpMethod', 'GET')
    params = event.get('queryStringParameters') or {}
    raw_body = event.get('body') or '{}'
    try:
        body = json.loads(raw_body) if isinstance(raw_body, str) and raw_body else {}
    except Exception:
        body = {}
    headers_in = {k.lower(): v for k, v in (event.get('headers') or {}).items()}
    token = headers_in.get('x-employee-token', '').strip()

    action = params.get('action') or body.get('action') or ''

    if action == 'ping':
        return _ok({'pong': True})

    emp = auth_employee(token)
    if not emp:
        return _err(401, 'Требуется авторизация')
    employee_name = emp.get('full_name') or 'staff'

    try:
        if method == 'GET':
            if action == 'export_csv':   return export_csv(params)
            if action == 'export_text':  return export_text(params)
            if action == 'export_xlsx':  return export_xlsx(params)
            return _err(400, f'Неизвестное действие: {action}')
        if method == 'POST':
            if action == 'parse_xlsx':   return parse_xlsx(body)
            if action == 'parse_csv':    return parse_csv(body)
            if action == 'import':       return import_items(body, employee_name)
            return _err(400, f'Неизвестное действие: {action}')
        return _err(405, f'Метод {method} не поддерживается')
    except Exception as e:
        return _err(500, f'Ошибка: {str(e)}')
